#coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.chart import BarChart, Series, Reference,BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Border,colors,Fill,Font,GradientFill,NamedStyle, Side,PatternFill,Protection
from datetime import datetime
import sys,time,locale,traceback,math,calendar
sys.path.append("..\\..")
from libreria.conexionBD import *

# ESTABLECE ESTILOS Y FORMATO A LASS CELDAS
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="000000")
nada = Side(border_style="none", color="000000")

border = Border(top=thin, left=thin, right=thin, bottom=thin)
fill = PatternFill(start_color='4db8ff',end_color='4db8ff', fill_type='solid')
font = Font(b=True, color="000000")
alignment = Alignment(horizontal="center", vertical="center")

class Ejecutor(object):
	def __init__(self,listaParametros,codSolic,rutaArchivo):
			#print "Metodo Constructor del Ejecutor Reporte Diario Total"
			try:
				libroExcel = Workbook()
				libroExcel.remove_sheet(libroExcel.get_sheet_by_name('Sheet'))

				logoRuta = "/home/settleradm/Aplicacion/reportes/libreria/img/LogoNetuno.png"
				#********************************************************************************************************
				# TABLA TEMPORAL
				query = "SELECT * FROM icx.ICX_RP_DR_TMP ORDER BY 'NB_OPERADORA'"
				query +=";"
				'''registros = ejecutarQuery(query)'''
				# TABLA CON LOS NOMBRES DE LAS OPERADORAS WHOLESALE Y SU EQUIVALENTE SUPPLIER
				query = "SELECT * FROM icx.ICX_RP_DR_ACDO_BILAT ORDER BY NB_OPERADORA_CLI;"
				supplierAsocClient = ejecutarQuery(query)
				# TABLA CON LOS NOMBRE DE LAS OPERADORAS Y SUS EQUIVALENTES ABREVIADOS
				query = "SELECT * FROM icx.ICX_RP_DR_ACDO_BILAT_PROV_DET;"
				nombCortos = ejecutarQuery(query)
				# TABLA CON LOS NOMBRE DE LAS OPERADORAS ABREVIADOS Y EL ORDEN EN QUE DEBEN ESTAR
				query = "SELECT * FROM icx.ICX_RP_DR_ACDO_BILAT_PROV ORDER BY I_ORDEN;"
				nombCortosOrden = ejecutarQuery(query)
				# EL QUERY OBTIENE LA FECHA DE LA TABLA DE SOLICITUD DE REPORTES
				query = "SELECT DATE(F_SOLIC) FROM icx.ICX_SOLIC_REPORTE WHERE C_REPORTE = 21;"
				fchRep = ejecutarQuery(query)
				fechaHoy = fchRep[0]['DATE(F_SOLIC)']
				# ATENCION SEBE REMOVER DESPUES DE LAS PRUEBAS
				#fechaHoy = datetime(fechaHoy.year,9,1)
				# ATENCION SEBE REMOVER DESPUES DE LAS PRUEBAS
				mes = ""
				fch = ""
				if fechaHoy.strftime('%d') == "01":
					if fechaHoy.strftime('%m')== "01":
						fch = datetime(fechaHoy.year - 1,12,fechaHoy.day)
					else:
						fch = datetime(fechaHoy.year,fechaHoy.month - 1,fechaHoy.day)
				else:
					fch = fechaHoy

				mes = fch.strftime('%m')
				ultDiaMes = calendar.monthrange(fch.year,fch.month)
				ultDiaMes = ultDiaMes[1]
				yymm = str(fch.year) + str(mes)
				#print yymm
				self.localtime = time.asctime(fch.timetuple())
				self.localtime = self.localtime.split(' ')
				self.localtime= str(self.localtime[1]) +' '+ str(self.localtime[4])

				#print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				#print len(registros)
				#print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				listWholesale = []
				listWholesaleGrupos = []
				listSupplier = []
				listSupplierWholesale = []
				listSupplierNotWholesale = []
				listRetail = []
				listDid = []
				listSuppWholesale = []
				listSuppRetail = []
				listBilatDest = [] # AGRUPANDO SEGÚN PROVEEDOR DE TERMINACIÓN
				listBilatDestDet = []
				listGraphReport = []
				#************************************************************************************************
				# SEPARACION DE LOS REGISTROS SEGUN NB_TIPO, EN DISTINTAS LISTAS
				for reg in ejecutarQuery_v2(50000,query):
					if reg['NB_TIPO'] == 'WHOLESALE':
						if reg['NB_GRUPO'] != 'NO ETELIX':
							listWholesaleGrupos.append(reg)
						else:
							listWholesale.append(reg)

					elif reg['NB_TIPO'] == 'SUPPLIER':
						listSupplier.append(reg)

					elif reg['NB_TIPO'] == 'RETAIL':
						listRetail.append(reg)

					elif reg['NB_TIPO'] == 'DID':
						listDid.append(reg)

					elif reg['NB_TIPO'] == 'SUPP WHOLESALE':
						listSuppWholesale.append(reg)

					elif reg['NB_TIPO'] == 'SUPP RETAIL':
						listSuppRetail.append(reg)

					elif reg['NB_TIPO'] == 'BILAT DEST':
						listBilatDest.append(reg)

					elif reg['NB_TIPO'] == 'BILAT DEST DET':
						listBilatDestDet.append(reg)

					elif reg['NB_TIPO'] == 'GRAPHIC REPORT':
						listGraphReport.append(reg)
				#************************************************************************************************
				# BLOQUE DE CODIGO QUE SUMA LOS CAMPOS DE TODOS LOS REGISTROS QUE NO PERTENECEN AL 'NB_GRUPO'= ' NO ETELIX'
				# EL RESULTADO ES UN SOLO DICCIONARIO PYTHON 'diccTemp' QUE SE AGREGA COMO APENDICE A LA LISTA
				# DE DICCIONARIOS listWholesale
				diccTemp = {}
				listGroupOper = []
				for i in  listWholesaleGrupos:
					if listGroupOper.count(i['NB_GRUPO']) ==0:
						listGroupOper.append(i['NB_GRUPO'])

				for nombOper in listGroupOper:
					diccTemp['NB_OPERADORA'] = 'WHL_'+nombOper
					for dicc in listWholesaleGrupos:
						if dicc['NB_GRUPO'] == nombOper:
							diccTemp['Q_CALL_SUM'] = diccTemp.get('Q_CALL_SUM',0) + dicc['Q_CALL_SUM']
							diccTemp['Q_MIN_SUM'] = diccTemp.get('Q_MIN_SUM',0) + dicc['Q_MIN_SUM']
							diccTemp['Q_COST_SUM'] = diccTemp.get('Q_COST_SUM',0) + dicc['Q_COST_SUM']
							diccTemp['Q_REV_SUM'] = diccTemp.get('Q_REV_SUM',0) + dicc['Q_REV_SUM']
							for i in range(1,32):
								diccTemp['Q_CALL_'+ str(i)] = diccTemp.get('Q_CALL_'+ str(i), 0) + dicc['Q_CALL_'+ str(i)]
								diccTemp['Q_MIN_'+ str(i)] = diccTemp.get('Q_MIN_'+ str(i), 0) + dicc['Q_MIN_'+ str(i)]
								diccTemp['Q_COST_'+ str(i)] = diccTemp.get('Q_COST_'+ str(i), 0) + dicc['Q_COST_'+ str(i)]
								diccTemp['Q_REV_'+ str(i)] = diccTemp.get('Q_REV_'+ str(i), 0) + dicc['Q_REV_'+ str(i)]
					listWholesale.append(diccTemp)
					diccTemp = {}
				listWholesaleGrupos = [] # SE LIMPIA DE LA MEMORIA YA NO SE NECESITA
				#************************************************************************************************
				# ESTE BLOQUE DE CODIGO DE BUSCA EN LA LA TUPLA DE DICCIONARIOS 'supplierAsocClient' CUALES NOMBRES
				# WHOLESALE POSEEN UN EQUIVALENTE NOMBRE SUPPLIER Y LUEGO BUSCAN EN LA LISTA DE DICCIONARIOS 'listSupplier'
				# CON EL NOMBRE SUPPLIER LOS DATOS QUE CORRESPONDEN AL NOMBRE WHOLSALE, ESTOS DATOS SE GUARDAN EN UN DICCIONARIO
				# LOS DATOS SE SUMAN SI EL WHOSALE TIENE MAS DE UN ASOCIADO SUPPLIER, EL DICCIONARIO SE GUANDA EN UNA
				# LISTA DE DICCIONARIOS 'listSupplierWholesale'
				diccTemp = {}
				for diccListWholesale in listWholesale:
					for diccSupplierAsocClient in supplierAsocClient:
						if diccListWholesale['NB_OPERADORA'] == diccSupplierAsocClient['NB_OPERADORA_CLI']:
							for diccListSupplier in listSupplier:
								if diccSupplierAsocClient['NB_OPERADORA_PROV'] ==  diccListSupplier['NB_OPERADORA']:
									diccTemp['NB_WHOSALE'] = diccListWholesale['NB_OPERADORA']
									diccTemp['Q_CALL_SUM'] = diccTemp.get('Q_CALL_SUM',0) + diccListSupplier['Q_CALL_SUM']
									diccTemp['Q_MIN_SUM'] = diccTemp.get('Q_MIN_SUM',0) + diccListSupplier['Q_MIN_SUM']
									diccTemp['Q_COST_SUM'] = diccTemp.get('Q_COST_SUM',0) + diccListSupplier['Q_COST_SUM']
									for i in range(1,32):
										diccTemp['Q_CALL_'+ str(i)] = diccTemp.get('Q_CALL_'+ str(i), 0) + diccListSupplier['Q_CALL_'+ str(i)]
										diccTemp['Q_MIN_'+ str(i)] = diccTemp.get('Q_MIN_'+ str(i), 0) + diccListSupplier['Q_MIN_'+ str(i)]
										diccTemp['Q_COST_'+ str(i)] = diccTemp.get('Q_COST_'+ str(i), 0) + diccListSupplier['Q_COST_'+ str(i)]
							if bool(diccTemp) == True:
								listSupplierWholesale.append(diccTemp)
								diccTemp = {}
				#************************************************************************************************
				# ESTE BLOQUE DE CODIGO BUSCA AQUELLOS SUPLLIER QUE NO POSEEN ASOCIACION CON CLIENTE O WHOLESALE
				# LOS AGREGA LA LISTA
				list_NB_OPERADORA_PROV = []
				for diccSupplierAsocClient in supplierAsocClient:
					list_NB_OPERADORA_PROV.append(diccSupplierAsocClient['NB_OPERADORA_PROV'])

				for diccListSupplier in listSupplier:
					if list_NB_OPERADORA_PROV.count(diccListSupplier['NB_OPERADORA']) == 0:
						listSupplierNotWholesale.append(diccListSupplier)
				# SE LIMPIAN DE LA MEMORIA YA NO SE NECESITAN
				supplierAsocClient = []
				list_NB_OPERADORA_PROV = []
				listSupplier = []
				#************************************************************************************************
				# EN ESTE BLOQUE TODAS LAS LISTAS QUE CONTINEN DICCIONARIOS CON LOS REGISTROS QUE VIENEN DE BASE DE DATOS
				# SE AGREGAN A UN DICCIONARIO 'diccDatos' PARA PODER PASAR Y UTILIZAR LOS DATOS EFICIENTEMENTE
				# DENTRO DEL RESTO DEL CODIGO
				diccDatos = {'listWholesale':listWholesale,
							'listSupplierWholesale':listSupplierWholesale,
							'listSupplierNotWholesale':listSupplierNotWholesale,
							'listRetail':listRetail,
							'listDid':listDid,
							'listSuppWholesale':listSuppWholesale,
							'listSuppRetail':listSuppRetail,
							'listBilatDest':listBilatDest,
							'nombCortos':nombCortos,
							'nombCortosOrden':nombCortosOrden,
							'listBilatDestDet':listBilatDestDet}
				#************************************************ HOJA EXCEL REPORTES DIARIOS ********************************
				for i in range(1,32):
					hojaExcel1 = libroExcel.create_sheet(title=str(i))
					self.ReportDiario(hojaExcel1,str(i),diccDatos,logoRuta)

				'''hojaExcel2 = libroExcel.create_sheet(title="01")
				self.ReportDiario(hojaExcel2,'1',diccDatos,logoRuta)'''
				#************************************************* HOJA EXCEL ACCUMULATED ************************************
				hojaExcelA = libroExcel.create_sheet(title="Accumulated")
				img = Image(logoRuta)
				hojaExcelA.add_image(img, 'K2')

				hojaExcelA.sheet_properties.tabColor = 'ff0000'
				hojaExcelA.sheet_view.showGridLines = False

				hojaExcelA.column_dimensions['A'].width = 1.5 + 0.5
				hojaExcelA.column_dimensions['B'].width = 2.0 + 0.5
				hojaExcelA.column_dimensions['C'].width = 1.57 + 0.5
				hojaExcelA.column_dimensions['D'].width = 36.0
				hojaExcelA.column_dimensions['E'].width = 18.71
				hojaExcelA.column_dimensions['F'].width = 12.5
				hojaExcelA.column_dimensions['G'].width = 12.5
				hojaExcelA.column_dimensions['H'].width = 12.5
				hojaExcelA.column_dimensions['I'].width = 0.58 + 0.5
				hojaExcelA.column_dimensions['J'].width = 12.5
				hojaExcelA.column_dimensions['K'].width = 1.0 + 0.5
				hojaExcelA.column_dimensions['L'].width = 12.5
				hojaExcelA.column_dimensions['M'].width = 12.0 + 0.5
				hojaExcelA.column_dimensions['N'].width = 12.0
				hojaExcelA.column_dimensions['O'].width = 1.5 + 0.5
				hojaExcelA.column_dimensions['P'].width = 12.43 + 0.5
				hojaExcelA.column_dimensions['Q'].width = 9.14 + 0.5
				hojaExcelA.column_dimensions['R'].width = 11.0
				hojaExcelA.column_dimensions['S'].width = 11.0

				hojaExcelA['B2'] = "DAILY REPORT"
				hojaExcelA['B2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')

				hojaExcelA['G2'] = "ACCUMULATED"
				hojaExcelA['G2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')
				hojaExcelA['G2'].alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelA['G3'] = self.localtime
				hojaExcelA['G3'].font = Font(color="044ea4", name='Arial')
				hojaExcelA['G3'].alignment = Alignment(horizontal="center", vertical="center")
				#******************************************************Information Resume*************************************
				filaInicio = self.InformationResume(hojaExcelA,'SUM',diccDatos,'044ea4')
				# filaInicio =  AL NUMERO DE LA FILA DONDE INICIARA EL SIGUIENTE CUADRO DE INFORMACION
				#*************************************************Client Information detail***********************************
				filaInicio = self.ClientInformationDetail(hojaExcelA,'SUM',filaInicio,diccDatos,'044ea4')
				#************************************************* Provider Information detail *******************************
				filaInicio = self.ProviderInformationdetail(hojaExcelA,'SUM',filaInicio,diccDatos,'044ea4') #26734d VERDE OSCURO
				#************************************ Bilaterals Destinations (minutes by provider)***************************
				filaInicio = self.BilateralsDestinations(hojaExcelA,'SUM',filaInicio,diccDatos)
				fila = filaInicio + 1

				hojaExcelA.merge_cells('F'+str(fila)+':G'+str(fila))
				hojaExcelA['F'+str(fila)].alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelA['F'+str(fila)] = "Wholesale"

				hojaExcelA.merge_cells('H'+str(fila)+':J'+str(fila))
				hojaExcelA['H'+str(fila)].alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelA['H'+str(fila)] = "Retail"

				hojaExcelA.merge_cells('L'+str(fila)+':M'+str(fila))
				hojaExcelA['L'+str(fila)].alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelA['L'+str(fila)] = "Total"

			 	for row in hojaExcelA.iter_cols(min_row=fila, max_row=fila, min_col=6, max_col=13):
				    for cell in row:
				    	if str(cell).find('.K') == -1:
					    	cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
					    	cell.border = Border(top=nada, left=nada, right=nada, bottom=nada)
					    	cell.alignment = Alignment(horizontal="center", vertical="center")

			 	for row in hojaExcelA.iter_cols(min_row=fila+1, max_row=fila+1, min_col=4, max_col=13):
				    for cell in row:
				    	if str(cell).find('.K') == -1:
					    	cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
					    	cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)
					    	cell.alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelA['D'+str(fila+1)] = "PROVIDER"
				hojaExcelA['E'+str(fila+1)] = "DESTINATIONS"
				hojaExcelA['F'+str(fila+1)] = "MIN"
				hojaExcelA['G'+str(fila+1)] = "COST"
				hojaExcelA['H'+str(fila+1)] = "MIN"
				hojaExcelA['J'+str(fila+1)] = "COST"
				hojaExcelA['L'+str(fila+1)] = "MIN"
				hojaExcelA['M'+str(fila+1)] = "COST"
				#*************************************************************************************************************
				#*************************************************************************************************************
				#*************************************************************************************************************
				# BLOQUE DE CODIGO QUE LLENA LA TABLA BILATERAL DESTINATION DET
				# 'listNB_PROV' ES UNA LISTA CON LOS NOMBRES COMPLETO DE LAS OPERADORAS
				listNB_PROV = []
				for diccNombCorto in nombCortos:
					listNB_PROV.append(diccNombCorto['NB_PROV'])
				# SE CREA UN NUEVO DICCIONARIO CON LOS DATOS DE listBilatDestDet PERO AHORA CON EL NOMBRE ABREVIADO DE LA OPERADORA
				listTemp = []
				diccTemp ={}
				for diccBilatDestDet in listBilatDestDet:
					if listNB_PROV.count(diccBilatDestDet['NB_OPERADORA']) > 0:
						for nombCortoOper in nombCortos:
							if nombCortoOper['NB_PROV'] == diccBilatDestDet['NB_OPERADORA']:
								diccTemp['NB_OPERADORA'] = nombCortoOper['NB_CORTO_PROV']
					else:
						diccTemp['NB_OPERADORA'] = 'OTHERS'
					diccTemp['NB_GRUPO'] = diccBilatDestDet['NB_GRUPO']
					diccTemp['Q_MIN_SUM'] = diccBilatDestDet['Q_MIN_SUM']
					diccTemp['Q_COST_SUM'] = diccBilatDestDet['Q_COST_SUM']
					diccTemp['Q_MIN_1'] = diccBilatDestDet['Q_MIN_1']
					diccTemp['Q_COST_1'] = diccBilatDestDet['Q_COST_1']
					listTemp.append(diccTemp)
					diccTemp ={}
				# 'listPaisOper' ES UNA LISTA CON LOS NOMBRES DE LOS PAISES Y EL NOMBRE DE SU OPERADORA
				listPaisOper = []
				nombreBuffer = ''
				for i in listTemp:
					if nombreBuffer.find(i['NB_OPERADORA'] + i['NB_GRUPO']) == -1:
						listPaisOper.append([i['NB_OPERADORA'],i['NB_GRUPO']])
						nombreBuffer += i['NB_OPERADORA'] + i['NB_GRUPO']
				# listNueva CONTENDRA LOS DICCIONARIOS DE listBilatDestDet PERO AHORA CON EL NOMBRE ABREVIADO DE LA OPERADORA
				listNueva = []
				for paisOper in listPaisOper:
					for diccTemp2 in listTemp:
						if paisOper[0] ==diccTemp2['NB_OPERADORA'] and paisOper[1]== diccTemp2['NB_GRUPO']:
							diccTemp['NB_OPERADORA'] = diccTemp2['NB_OPERADORA']
							diccTemp['NB_GRUPO'] = diccTemp2['NB_GRUPO']
							diccTemp['Q_MIN_SUM'] = diccTemp.get('Q_MIN_SUM',0)+diccTemp2['Q_MIN_SUM']
							diccTemp['Q_COST_SUM'] = diccTemp.get('Q_COST_SUM',0)+diccTemp2['Q_COST_SUM']
							diccTemp['Q_MIN_1'] = diccTemp.get('Q_MIN_1',0)+diccTemp2['Q_MIN_1']
							diccTemp['Q_COST_1'] = diccTemp.get('Q_COST_1',0)+diccTemp2['Q_COST_1']
					listNueva.append(diccTemp)
					diccTemp={}
				listBilatDestDet = listNueva
				listNueva = []
				listTemp = []
				diccTemp={}
				# SE SEPARAN DE LA LISTA listBilatDestDet LOS REGISTROS CON DESTINO PAIS IGUAL A OTHERS 'NB_GRUPO'=='OTHERS'
				for i in listBilatDestDet:
					if i['NB_GRUPO']=='OTHERS':
						listTemp.append(i)
					else:
						listNueva.append(i)
				listBilatDestDet = listNueva
				# EN LA LISTA listTemp SE ENCUENTRA TODOS LOS REGISTROS CON 'NB_GRUPO'=='OTHERS', ENTONCES SE BUSCA EN
				# LA LISTA listNueva CON LOS REGISTROS 'NB_GRUPO'!='OTHERS' AQUELLOS QUE CONICINDAN EN EN
				# EL CAMPO NB_OPERADORA Y SE SUMAN A LOS DICCIONARIOS DE listTemp PERO SON AGREGADOS A listBilatDestDet
				# PARA TENER LA LISTA DE DICCIONARIOS CON TODOS LOS REGISTROS DE TODOSS LOS NB_GRUPO
				for diccTemp in listTemp:
					for diccBilatDet in listNueva:
						if diccBilatDet['NB_OPERADORA'] == diccTemp['NB_OPERADORA']:
							diccTemp['NB_OPERADORA'] = diccBilatDet['NB_OPERADORA']
							diccTemp['NB_GRUPO'] = 'OTHERS'
							diccTemp['Q_MIN_SUM'] = diccTemp.get('Q_MIN_SUM',0)+diccBilatDet['Q_MIN_SUM']
							diccTemp['Q_COST_SUM'] = diccTemp.get('Q_COST_SUM',0)+diccBilatDet['Q_COST_SUM']
							diccTemp['Q_MIN_1'] = diccTemp.get('Q_MIN_1',0)+diccBilatDet['Q_MIN_1']
							diccTemp['Q_COST_1'] = diccTemp.get('Q_COST_1',0)+diccBilatDet['Q_COST_1']
					listBilatDestDet.append(diccTemp)
					diccTemp={}
				##############################################################################
				sumTotal = [0]*6
				listTemp = []
				for diccBilatDestDet in listBilatDestDet:
					totalMin = math.fabs(diccBilatDestDet['Q_MIN_SUM']) + math.fabs(diccBilatDestDet['Q_MIN_1'])
					totalCost = math.fabs(diccBilatDestDet['Q_COST_SUM']) + math.fabs(diccBilatDestDet['Q_COST_1'])
					sumTotal[0] = sumTotal[0] + math.fabs(diccBilatDestDet['Q_MIN_SUM'])
					sumTotal[1] = sumTotal[1] + math.fabs(diccBilatDestDet['Q_COST_SUM'])
					sumTotal[2] = sumTotal[2] + math.fabs(diccBilatDestDet['Q_MIN_1'])
					sumTotal[3] = sumTotal[3] + math.fabs(diccBilatDestDet['Q_COST_1'])
					sumTotal[4] = sumTotal[4] + totalMin
					sumTotal[5] = sumTotal[5] + totalCost
					listTemp.append([totalMin,totalCost,diccBilatDestDet])
				listTemp.sort()
				listTemp.reverse()
				#**************************************************************************************************************
				n=fila+2
				nf = n
				fila = n-1+ len(listBilatDestDet)
				while (n<= fila):
				 	for row in hojaExcelA.iter_cols(min_row= n, max_row= n, min_col=4, max_col=13):
					    for cell in row:
							cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
					n = n + 2
				#***************************************************************************************************************
				for diccBilatDestDet in listTemp:
					hojaExcelA.cell(row=nf, column=4, value=diccBilatDestDet[2]['NB_OPERADORA'])
					hojaExcelA.cell(row=nf, column=5, value=diccBilatDestDet[2]['NB_GRUPO'])
					hojaExcelA.cell(row=nf, column=6, value=math.fabs(diccBilatDestDet[2]['Q_MIN_SUM'])).number_format = '#,##0.00'
					hojaExcelA.cell(row=nf, column=7, value=math.fabs(diccBilatDestDet[2]['Q_COST_SUM'])).number_format = '#,##0.00'
					hojaExcelA.cell(row=nf, column=8, value=math.fabs(diccBilatDestDet[2]['Q_MIN_1'])).number_format = '#,##0.00'
					hojaExcelA.cell(row=nf, column=10, value=math.fabs(diccBilatDestDet[2]['Q_COST_1'])).number_format = '#,##0.00'
					hojaExcelA.cell(row=nf, column=12, value=diccBilatDestDet[0]).number_format = '#,##0.00'
					hojaExcelA.cell(row=nf, column=13, value=diccBilatDestDet[1]).number_format = '#,##0.00'
					nf = nf +1

				hojaExcelA.cell(row=nf, column=6, value=sumTotal[0]).number_format = '#,##0.00'
				hojaExcelA.cell(row=nf, column=7, value=sumTotal[1]).number_format = '#,##0.00'
				hojaExcelA.cell(row=nf, column=8, value=sumTotal[2]).number_format = '#,##0.00'
				hojaExcelA.cell(row=nf, column=10, value=sumTotal[3]).number_format = '#,##0.00'
				hojaExcelA.cell(row=nf, column=12, value=sumTotal[4]).number_format = '#,##0.00'
				hojaExcelA.cell(row=nf, column=13, value=sumTotal[5]).number_format = '#,##0.00'
				# SE LIMPIA DE LA MEMORIA
				listTemp = []
				listNueva = []
				listPaisOper = []
				diccTemp ={}
				#***************************************************************************************************************
				#***************************************************************************************************************
				#***************************************************************************************************************
				hojaExcelA['D'+str(fila+1)] = 'Total'
				hojaExcelA['D'+str(fila+1)].alignment = Alignment(horizontal="right", vertical="center")
				hojaExcelA['D'+str(fila+1)].font = Font(b=True, name='Arial')
			 	for row in hojaExcelA.iter_cols(min_row=fila+1, max_row=fila+1, min_col=5, max_col=13):
				    for cell in row:
						cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
				#************************************************ HOJA EXCEL ACCUMULATED REV BY DATE ***************************
				hojaExcelARD = libroExcel.create_sheet(title="Accumulated Rev by date")
				tipoAcumulado = "Accumulated by date - Revenues"
				AccumBy = "Accum Revenue"
				self.AccumulatedBy(hojaExcelARD,tipoAcumulado,AccumBy,diccDatos,logoRuta)
				#********************************************** HOJA EXCEL ACCUMULATED MIN BY DATE *****************************
				hojaExcelAMD = libroExcel.create_sheet(title="Accumulated Min by date")
				tipoAcumulado = "Accumulated by date - Minutes"
				AccumBy = "Accum Minutes"
				self.AccumulatedBy(hojaExcelAMD,tipoAcumulado,AccumBy,diccDatos,logoRuta)
				diccDatos = {}
				#******************************************** HOJA EXCEL GRAPHIC REPORT ***************************************
				hojaExcelGR = libroExcel.create_sheet(title="Graphic Report")

				img = Image(logoRuta)
				hojaExcelGR.add_image(img, 'K2')

				hojaExcelGR.sheet_properties.tabColor = 'FFFF00'
				hojaExcelGR.sheet_view.showGridLines = False

				hojaExcelGR.column_dimensions['A'].width = 1.5
				hojaExcelGR.column_dimensions['B'].width = 15.0
				hojaExcelGR.column_dimensions['C'].width = 14.0
				hojaExcelGR.column_dimensions['D'].width = 14.0
				hojaExcelGR.column_dimensions['E'].width = 14.0
				hojaExcelGR.column_dimensions['F'].width = 14.0
				hojaExcelGR.column_dimensions['G'].width = 1.5
				hojaExcelGR.column_dimensions['H'].width = 10.0
				hojaExcelGR.column_dimensions['I'].width = 10.0
				hojaExcelGR.column_dimensions['J'].width = 10.0
				hojaExcelGR.column_dimensions['K'].width = 10.0
				hojaExcelGR.column_dimensions['L'].width = 10.0
				hojaExcelGR.column_dimensions['M'].width = 1.5
				hojaExcelGR.column_dimensions['N'].width = 10.0
				hojaExcelGR.column_dimensions['O'].width = 10.0
				hojaExcelGR.column_dimensions['P'].width = 10.0
				hojaExcelGR.column_dimensions['Q'].width = 10.0
				hojaExcelGR.column_dimensions['R'].width = 1.5
				hojaExcelGR.column_dimensions['S'].width = 10.0

				hojaExcelGR['B2'] = "DAILY REPORT"
				hojaExcelGR['B2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')

				hojaExcelGR['G2'] = "ACCUMULATED"
				hojaExcelGR['G2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')
				hojaExcelGR['G2'].alignment = Alignment(horizontal="center", vertical="center")
				hojaExcelGR['G3'] = self.localtime
				hojaExcelGR['G3'].font = Font(color="044ea4", name='Arial')
				hojaExcelGR['G3'].alignment = Alignment(horizontal="center", vertical="center")

				hojaExcelGR['A5'].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
				hojaExcelGR['B5'] = "► Statistic Graphic Report"
				for row in hojaExcelGR.iter_cols(min_row=5, max_row=5, min_col=2, max_col=17):
				    for cell in row:
						cell.fill = PatternFill(start_color="044ea4",end_color="044ea4", fill_type='solid') # COLOR AZUL
						cell.font = Font(b=True, color="ffffff",size = "10", name='Arial')

				hojaExcelGR['B7'] = "► Wholesale - Venezuela - Cellular traffic by Provider"
				hojaExcelGR['B7'].font = Font(color="044ea4", name='Arial')

				hojaExcelGR.row_dimensions[8].height = 3.0
				cantOper = len(listGraphReport)
			 	for row in hojaExcelGR.iter_cols(min_row=9, max_row=9, min_col=3, max_col=3+cantOper):
				    for cell in row:
						cell.fill = PatternFill(start_color='FFFF00',end_color='FFFF00', fill_type='solid') # COLOR AMARILLO
				 		cell.font = Font(b=True, color="000000", name='Arial')

			 	for row in hojaExcelGR.iter_cols(min_row=9, max_row=40, min_col=2, max_col=2+cantOper +1):
				    for cell in row:
						cell.alignment = Alignment(horizontal="right", vertical="center")
						cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

				for row in hojaExcelGR.iter_cols(min_row=41, max_row=41, min_col=3, max_col=2+cantOper +1):
				    for cell in row:
						cell.alignment = Alignment(horizontal="right", vertical="center")
						cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

			 	for row in hojaExcelGR.iter_cols(min_row=9, max_row=40, min_col=2+cantOper+1, max_col=2+cantOper +1):
				    for cell in row:
						cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
				 		cell.font = Font(color="000000", name='Arial')

				n = 1
				for row in hojaExcelGR.iter_cols(min_row=10, max_row=40, min_col=2, max_col=2):
				    for cell in row:
				    	cell.value = n
				    	cell.alignment = Alignment(horizontal="center", vertical="center")
				    	n = n +1
				#**********************************************************************************************
				#**********************************************************************************************
				#**********************************************************************************************
				diccGraphRepot = listGraphReport
				totalOper = [0]*cantOper
				hojaExcelGR['B9'] = "DIA"
				hojaExcelGR['B9'].alignment = Alignment(horizontal="center", vertical="center")
				for i in range(0,cantOper):
					hojaExcelGR.cell(row=9, column=3+i, value=diccGraphRepot[i]['NB_OPERADORA'])
				hojaExcelGR.cell(row=9, column=2+cantOper+1, value="Total Minutes")

				totalOper = 0
				totalFinalOper = 0
				totalMesOper = [0]*cantOper
				for i in range(1,32):
					for j in range(0,cantOper):
						hojaExcelGR.cell(row=9+i, column=3+j, value=diccGraphRepot[j]['Q_MIN_'+ str(i)]).number_format = '#,##0.00'
						totalMesOper[j] = totalMesOper[j] + diccGraphRepot[j]['Q_MIN_'+ str(i)]
						totalOper = totalOper + diccGraphRepot[j]['Q_MIN_'+ str(i)]
					hojaExcelGR.cell(row=9+i, column=2+cantOper+1, value=totalOper).number_format = '#,##0.00'
					totalFinalOper = totalFinalOper + totalOper
					totalOper = 0

				for j in range(0,cantOper):
					hojaExcelGR.cell(row=41, column=3+j, value=totalMesOper[j]).number_format = '#,##0.00'
				hojaExcelGR.cell(row=41, column=3+j+1, value=totalFinalOper).number_format = '#,##0.00'

				Bar = Reference(hojaExcelGR, min_col= 2, min_row= 10, max_row= 40)
				Data = Reference(hojaExcelGR, min_col= 3, max_col= 2+cantOper, min_row= 9 , max_row= 40)
				self.funcionGrafica(hojaExcelGR,2,'H9','Venezuela - Cellular traffic by Provider',Data,Bar)

				Data = Reference(hojaExcelGR, min_col= 3, max_col= 3, min_row= 9 , max_row= 40)
				self.funcionGrafica(hojaExcelGR,2,'H25','Venezuela - Cellular traffic - Digitel',Data,Bar)

				Data = Reference(hojaExcelGR, min_col= 4, max_col= 4, min_row= 9 , max_row= 40)
				self.funcionGrafica(hojaExcelGR,4,'H41','Venezuela - Cellular traffic - Movilnet',Data,Bar)

				Data = Reference(hojaExcelGR, min_col= 5, max_col= 5, min_row= 9 , max_row= 40)
				self.funcionGrafica(hojaExcelGR,5,'H57','Venezuela - Cellular traffic - Movistar',Data,Bar)
				#**********************************************************************************************
				#**********************************************************************************************
				#**********************************************************************************************
				#***************************************************** HOJA EXCEL *****************************************
				#nombreReporte = "DR"+fch.strftime('%Y%m%d')+".xlsx"
				fecha = datetime.now().strftime('%Y%m%d-%H%M%S')
				nombreReporte = "DR"+fecha+".xlsx"
				libroExcel.save(rutaArchivo + nombreReporte)
				self.rutaArchivo = rutaArchivo + nombreReporte
				self.contenido = listaParametros["cuerpoCorreo"].replace("[/NB_ARCHIVO]",nombreReporte)
				cuadroResumen ='''
				<p>Resumen al fechaHoy</p>
				<table style="border: 1px solid black;border-collapse: collapse;">
				  <tr style="background-color: #4f81bd;color: white;">
				    <th style="border: 1px solid black;">Tipo</th>
				    <th style="border: 1px solid black;">fechaAyer</th>
				    <th style="border: 1px solid black;">fechaHoy</th>
				    <th style="border: 1px solid black;">Total</th>
				    <th style="border: 1px solid black;">Promedio</th>
				  </tr>
				  <tr style="background-color: #dbe5f1">
				    <td style="border: 1px solid black;">REVENUE WHL</td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				  </tr>
				  <tr>
				    <td style="border: 1px solid black;">COSTO WHL</td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				  </tr>
				  <tr style="background-color: #dbe5f1">
				    <td style="border: 1px solid black;">COSTO RET (No Panama)</td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				  </tr>
				  <tr>
				    <td style="border: 1px solid black;">COSTO RET (Panama)</td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				    <td style="border: 1px solid black;">USD </td>
				  </tr>
				</table>
				<p>Cost WHL/RET: VITCOM no es incluido</p>
				'''
				self.contenido += cuadroResumen
				self.asunto = listaParametros["tituloCorreo"]

				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '{0}', F_ULT_ACT = NOW() WHERE C_SOLICITUD = {1}"
				ejecutarQuery(query.format("Ruta del archivo: " + self.rutaArchivo , codSolic))

			except Exception as ex:
				detalles = traceback.format_exc()
				observacion = "Excepcion de tipo {0} . Argumentos:\n{1!r}\nDetalles:\n{2} "
				observacion = observacion.format(type(ex).__name__, ex.args,detalles)
				observacion = observacion.replace("'","*")
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '"+ str(observacion)
				query += "',F_ULT_ACT = NOW() WHERE C_SOLICITUD = "+ str(codSolic) +" AND X_OBS IS NULL;"
				respuesta = ejecutarQuery(query)
	#*********************************************************************************************************************
	def InformationResume(self,hojaExcel,nombreHoja,Datos,colorFranja):
		hojaExcel['A5'].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['B5'] = "►"
		hojaExcel['C5'] = "Information Resume"
		for row in hojaExcel.iter_cols(min_row=5, max_row=5, min_col=2, max_col=16):
		    for cell in row:
		    	cell.fill = PatternFill(start_color=colorFranja,end_color=colorFranja, fill_type='solid') # COLOR AZUL
		    	cell.font = Font(b=True, color="ffffff",size = "10", name='Arial')

		hojaExcel.row_dimensions[6].height = 3.0
		hojaExcel.merge_cells('L7:N7')
		hojaExcel['L7'].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['L7'].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['L7'] = "Wholesale-Retail"

		hojaExcel['P7'].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['P7'].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['P7'] = "Revenue vs"

		hojaExcel['D8'] = "CLIENTS"
		hojaExcel['E8'] = "Calls"
		hojaExcel['F8'] = "Minutes"
		hojaExcel['G8'] = "Cost"
		hojaExcel['H8'] = "Revenue"
		hojaExcel['J8'] = "Profit"
		hojaExcel['L8'] = "Calls"
		hojaExcel['M8'] = "Minutes"
		hojaExcel['N8'] = "Cost"
		hojaExcel['P8'] = "Cost"
	 	for row in hojaExcel.iter_cols(min_row=8, max_row=8, min_col=4, max_col=16):
		    for cell in row:
		    	if str(cell).find('.I') == -1 and str(cell).find('.K') == -1 and str(cell).find('.O') == -1 :
			    	cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
			    	cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		hojaExcel['D9'] = '► Wholesale'
		hojaExcel['D10'] = '► Retail'
		hojaExcel['D11'] = 'Total'
		hojaExcel['D11'].alignment = Alignment(horizontal="right", vertical="center")
		hojaExcel['D11'].font = Font(b=True, name='Arial')
		hojaExcel['D13'] = '► DIDs'
		hojaExcel['E11'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
		hojaExcel['F11'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
		hojaExcel['G11'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
	 	hojaExcel['E14'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
		hojaExcel['F14'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
		hojaExcel['G14'].border = Border(top=thin, left=nada, right=nada, bottom=nada)
		#********************************************************************************************************
		#********************************************************************************************************
		#********************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS DEL CUADRO INFORMATION RESUME
		diccResum = {}
		diccTemp = {}
		for reg in Datos['listWholesale']:
			diccTemp['Q_CALL_'+ nombreHoja] = diccTemp.get('Q_CALL_'+ nombreHoja, 0) + reg['Q_CALL_'+ nombreHoja]
			diccTemp['Q_MIN_'+ nombreHoja] = diccTemp.get('Q_MIN_'+ nombreHoja, 0) + reg['Q_MIN_'+ nombreHoja]
			diccTemp['Q_COST_'+ nombreHoja] = diccTemp.get('Q_COST_'+ nombreHoja, 0) + reg['Q_COST_'+ nombreHoja]
			diccTemp['Q_REV_'+ nombreHoja] = diccTemp.get('Q_REV_'+ nombreHoja, 0) + reg['Q_REV_'+ nombreHoja]
		diccResum['diccWT'] = diccTemp
		diccTemp = {}
		for reg in Datos['listSupplierWholesale']:
			diccTemp['Q_CALL_'+ nombreHoja] = diccTemp.get('Q_CALL_'+ nombreHoja, 0) + reg['Q_CALL_'+ nombreHoja]
			diccTemp['Q_MIN_'+ nombreHoja] = diccTemp.get('Q_MIN_'+ nombreHoja, 0) + reg['Q_MIN_'+ nombreHoja]
			diccTemp['Q_COST_'+ nombreHoja] = diccTemp.get('Q_COST_'+ nombreHoja, 0) + reg['Q_COST_'+ nombreHoja]

		for reg in Datos['listSupplierNotWholesale']:
			diccTemp['Q_CALL_'+ nombreHoja] = diccTemp.get('Q_CALL_'+ nombreHoja, 0) + reg['Q_CALL_'+ nombreHoja]
			diccTemp['Q_MIN_'+ nombreHoja] = diccTemp.get('Q_MIN_'+ nombreHoja, 0) + reg['Q_MIN_'+ nombreHoja]
			diccTemp['Q_COST_'+ nombreHoja] = diccTemp.get('Q_COST_'+ nombreHoja, 0) + reg['Q_COST_'+ nombreHoja]
		diccResum['diccSWT'] = diccTemp
		diccTemp = {}
		for reg in Datos['listRetail']:
			diccTemp['Q_CALL_'+ nombreHoja] = diccTemp.get('Q_CALL_'+ nombreHoja, 0) + reg['Q_CALL_'+ nombreHoja]
			diccTemp['Q_MIN_'+ nombreHoja] = diccTemp.get('Q_MIN_'+ nombreHoja, 0) + reg['Q_MIN_'+ nombreHoja]
			diccTemp['Q_COST_'+ nombreHoja] = diccTemp.get('Q_COST_'+ nombreHoja, 0) + reg['Q_COST_'+ nombreHoja]
		diccResum['diccRT'] = diccTemp
		diccTemp = {}
		for reg in Datos['listDid']:
			diccTemp['Q_CALL_'+ nombreHoja] = diccTemp.get('Q_CALL_'+ nombreHoja, 0) + reg['Q_CALL_'+ nombreHoja]
			diccTemp['Q_MIN_'+ nombreHoja] = diccTemp.get('Q_MIN_'+ nombreHoja, 0) + reg['Q_MIN_'+ nombreHoja]
			diccTemp['Q_COST_'+ nombreHoja] = diccTemp.get('Q_COST_'+ nombreHoja, 0) + reg['Q_COST_'+ nombreHoja]
		diccResum['diccDT'] = diccTemp
		diccTemp = {}
		# WHOLESALE RESUMEN
		hojaExcel.cell(row=9, column=5, value= diccResum['diccWT']['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
		hojaExcel.cell(row=9, column=6, value= diccResum['diccWT']['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
		hojaExcel.cell(row=9, column=7, value= diccResum['diccWT']['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
		hojaExcel.cell(row=9, column=8, value= diccResum['diccWT']['Q_REV_'+ nombreHoja]).number_format = '#,##0.00'
		profitTotal = diccResum['diccWT']['Q_REV_'+ nombreHoja] - diccResum['diccWT']['Q_COST_'+ nombreHoja]
		if profitTotal < 0 :
			hojaExcel.cell(row=9, column=10).font = Font(color= 'ff0000')
			profitTotal = profitTotal * -1
		hojaExcel.cell(row=9, column=10, value= profitTotal ).number_format = '#,##0.00'
		# WHOLESALE-RETAIL RESUMEN
		hojaExcel.cell(row=9, column=12, value= diccResum['diccSWT']['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
		hojaExcel.cell(row=9, column=13, value= diccResum['diccSWT']['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
		hojaExcel.cell(row=9, column=14, value= diccResum['diccSWT']['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
		profitTotal = diccResum['diccWT']['Q_REV_'+ nombreHoja] - diccResum['diccSWT']['Q_COST_'+ nombreHoja]
		if profitTotal < 0 :
			hojaExcel.cell(row=9, column=16).font = Font(color= 'ff0000')
			profitTotal = profitTotal * -1
		hojaExcel.cell(row=9, column=16, value= profitTotal).number_format = '#,##0.00'
		# RETAIL RESUMEN
		hojaExcel.cell(row=10, column=5, value= diccResum['diccRT']['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
		hojaExcel.cell(row=10, column=6, value= diccResum['diccRT']['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
		hojaExcel.cell(row=10, column=7, value= diccResum['diccRT']['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
		# DID RESUMEN
		hojaExcel.cell(row=13, column=5, value= diccResum['diccDT']['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
		hojaExcel.cell(row=13, column=6, value= diccResum['diccDT']['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
		hojaExcel.cell(row=13, column=7, value= diccResum['diccDT']['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
		# TOTAL RESUMEN
		callTotal = 0
		minTotal = 0
		costTotal = 0
		diccResum.pop('diccSWT',0)
		diccResum.pop('diccDT',0)
		for i in diccResum:
			callTotal = callTotal + diccResum[i]['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + diccResum[i]['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + diccResum[i]['Q_COST_'+ nombreHoja]

		diccResum = {}
		hojaExcel.cell(row=11, column=5, value= callTotal).number_format = '#,##0'
		hojaExcel.cell(row=11, column=6, value= minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=11, column=7, value= costTotal).number_format = '#,##0.00'

		for row in hojaExcel.iter_cols(min_row=8, max_row=8, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")
		#********************************************************************************************************
		#********************************************************************************************************
		#********************************************************************************************************
		return 15 # ES EL NUMERO DE LA FILA DONDE INICIA EL SIGUIENTE CUADRO 'ClientInformationDetail'
	#********************************************************************************************************************
	def ClientInformationDetail(self,hojaExcel,nombreHoja,filaInicio,Datos,colorFranja):
		fila = filaInicio
		hojaExcel['A'+str(fila)].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['B'+str(fila)] = "►"
		hojaExcel['C'+str(fila)] = "Client Information detail"
		for row in hojaExcel.iter_cols(min_row=fila, max_row=fila, min_col=2, max_col=16):
		    for cell in row:
		    	cell.fill = PatternFill(start_color=colorFranja,end_color=colorFranja, fill_type='solid') # COLOR AZUL
		    	cell.font = Font(b=True, color="ffffff",size = "10", name='Arial')
		hojaExcel['C'+str(fila+2)] = "Wholesale"
		hojaExcel['C'+str(fila+2)].font = Font(b=True,color= '000000')

		hojaExcel.row_dimensions[fila+1].height = 3.0
		hojaExcel.merge_cells('L'+str(fila+2)+':N'+str(fila+2))
		hojaExcel['L'+str(fila+2)].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['L'+str(fila+2)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['L'+str(fila+2)] = "Wholesale-Retail"

		hojaExcel['P'+str(fila+2)].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['P'+str(fila+2)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['P'+str(fila+2)] = "Revenue vs"

		hojaExcel['D'+str(fila+3)] = "CLIENTS"
		hojaExcel['E'+str(fila+3)] = "Calls"
		hojaExcel['F'+str(fila+3)] = "Minutes"
		hojaExcel['G'+str(fila+3)] = "Cost"
		hojaExcel['H'+str(fila+3)] = "Revenue"
		hojaExcel['J'+str(fila+3)] = "Profit"
		hojaExcel['L'+str(fila+3)] = "Calls"
		hojaExcel['M'+str(fila+3)] = "Minutes"
		hojaExcel['N'+str(fila+3)] = "Cost"
		hojaExcel['P'+str(fila+3)] = "Cost"
	 	for row in hojaExcel.iter_cols(min_row=fila+3, max_row=fila+3, min_col=4, max_col=16):
		    for cell in row:
				if str(cell).find('.I') == -1 and str(cell).find('.K') == -1 and str(cell).find('.O') == -1 :
					cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
					cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)
		n=fila+4
		nf = n #!!!!!!!!!!!!!!!!!
		fila = fila + 4 -1 + len(Datos['listWholesale']) #len('listWholesale') CANTIDAD DE REGISTROS EN DICCIONARIO
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row= n, max_row= n, min_col=4, max_col=16):
				for cell in row:
					if str(cell).find('.I') == -1 and str(cell).find('.K') == -1 and str(cell).find('.O') == -1 :
						cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2

		for row in hojaExcel.iter_cols(min_row=nf-1, max_row=nf-1, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

	 	for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=5, max_col=16):
		    for cell in row:
				if str(cell).find('.I') == -1 and str(cell).find('.K') == -1 and str(cell).find('.O') == -1 :
					cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO WHOLESALE DE 'Client Information detail'
		callTotal = 0
		minTotal = 0
		costTotal = 0
		revTotal = 0
		callTotalwr = 0
		minTotalwr = 0
		costTotalwr = 0
		Datos['listWholesale'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listWholesale'])
		for regWholesal in Datos['listWholesale']:
			callTotal = callTotal + regWholesal['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + regWholesal['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + regWholesal['Q_COST_'+ nombreHoja]
			revTotal = revTotal + regWholesal['Q_REV_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=4, value=regWholesal['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=5, value=regWholesal['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=6, value=regWholesal['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value=regWholesal['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=8, value=regWholesal['Q_REV_'+ nombreHoja]).number_format = '#,##0.00'
			profit = regWholesal['Q_REV_'+ nombreHoja] - regWholesal['Q_COST_'+ nombreHoja]
			if profit < 0 :
				hojaExcel.cell(row=nf, column=10).font = Font(color= 'ff0000')
				profit = profit * -1
			hojaExcel.cell(row=nf, column=10, value= profit).number_format = '#,##0.00'
			# CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO WHOLESALE-RETAIL DE 'Client Information detail'
			for lsw in Datos['listSupplierWholesale']:
				if regWholesal['NB_OPERADORA'] == lsw['NB_WHOSALE']:
					hojaExcel.cell(row=nf, column=12, value=lsw['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
					hojaExcel.cell(row=nf, column=13, value=lsw['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
					hojaExcel.cell(row=nf, column=14, value=lsw['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
					revVsCost = regWholesal['Q_REV_'+ nombreHoja] - lsw['Q_COST_'+ nombreHoja]
					if revVsCost < 0 :
						hojaExcel.cell(row=nf, column=16).font = Font(color= 'ff0000')
						revVsCost = revVsCost * -1
					hojaExcel.cell(row=nf, column=16, value=revVsCost).number_format = '#,##0.00'
					break
				else:
					hojaExcel.cell(row=nf, column=12, value=0).number_format = '#,##0'
					hojaExcel.cell(row=nf, column=13, value=0).number_format = '#,##0.00'
					hojaExcel.cell(row=nf, column=14, value=0).number_format = '#,##0.00'
					hojaExcel.cell(row=nf, column=16, value=0).number_format = '#,##0.00'
			nf = nf +1
		# SE LLENAN LAS CELDAS CON LA SUMA TOTAL DE LOS VALORES DEL CUADRO WHOLESALE
		hojaExcel.cell(row=fila+1, column=5, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=fila+1, column=6, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=fila+1, column=7, value=costTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=fila+1, column=8, value=revTotal).number_format = '#,##0.00'
		profit = revTotal - costTotal
		if profit < 0 :
			hojaExcel.cell(row=fila+1, column=10).font = Font(color= 'ff0000')
			profit = profit * -1
		hojaExcel.cell(row=fila+1, column=10, value= profit).number_format = '#,##0.00'
		# SE LLENAN LAS CELDAS CON LA SUMA TOTAL DE LOS VALORES DEL CUADRO WHOLESALE-RETAIL
		for lsw in Datos['listSupplierWholesale']:
			callTotalwr = callTotalwr + lsw['Q_CALL_'+ nombreHoja]
			minTotalwr = minTotalwr + lsw['Q_MIN_'+ nombreHoja]
			costTotalwr = costTotalwr + lsw['Q_COST_'+ nombreHoja]

		hojaExcel.cell(row=fila+1, column=12, value=callTotalwr).number_format = '#,##0'
		hojaExcel.cell(row=fila+1, column=13, value=minTotalwr).number_format = '#,##0.00'
		hojaExcel.cell(row=fila+1, column=14, value=costTotalwr).number_format = '#,##0.00'
		profit = revTotal - costTotalwr
		if profit < 0 :
			hojaExcel.cell(row=fila+1, column=16).font = Font(color= 'ff0000')
			profit = profit * -1
		hojaExcel.cell(row=fila+1, column=16, value= profit).number_format = '#,##0.00'
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		fila = fila + 3
		hojaExcel['L'+str(fila)] = "Additional Information for"
		hojaExcel['L'+str(fila+1)] = "Net Cash control only"
		hojaExcel['L'+str(fila+2)] = "VITCOM & DID are not included"
		hojaExcel.row_dimensions[fila+3].height = 3.0

		hojaExcel['J'+str(fila+4)] = "Providers"
		hojaExcel['J'+str(fila+4)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['J'+str(fila+4)].border = Border(top=nada, left=nada, right=nada, bottom=thin)
		hojaExcel['L'+str(fila+4)] = "Calls"
		hojaExcel['L'+str(fila+4)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['L'+str(fila+4)].border = Border(top=nada, left=nada, right=nada, bottom=thin)
		hojaExcel['M'+str(fila+4)] = "Minutes"
		hojaExcel['M'+str(fila+4)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['M'+str(fila+4)].border = Border(top=nada, left=nada, right=nada, bottom=thin)
		hojaExcel['N'+str(fila+4)] = "Amount"
		hojaExcel['N'+str(fila+4)].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['N'+str(fila+4)].border = Border(top=nada, left=nada, right=nada, bottom=thin)

		fila = fila + 5
		filaMax = fila-1+ len(Datos['listSupplierNotWholesale']) #len('listSupplierNotWholesale') CANTIDAD DE REGISTROS EN DICCIONARIO
		for row in hojaExcel.iter_cols(min_row=fila, max_row=filaMax+2, min_col=12, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="right", vertical="center")
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO DE 'Additional Information for'
		nf = fila
		callTotal = 0
		minTotal = 0
		costTotal = 0
		revTotal = 0
		Datos['listSupplierNotWholesale'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listSupplierNotWholesale'])
		for lsnotw in Datos['listSupplierNotWholesale']:
			callTotal = callTotal + lsnotw['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + lsnotw['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + lsnotw['Q_COST_'+ nombreHoja]
			revTotal = revTotal + lsnotw['Q_REV_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=10, value=lsnotw['NB_OPERADORA']).alignment = Alignment(horizontal="right")
			hojaExcel.cell(row=nf, column=12, value=lsnotw['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=13, value=lsnotw['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=14, value=lsnotw['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			'''profit = lsnotw['Q_REV_'+ nombreHoja] - lsnotw['Q_COST_'+ nombreHoja]
			if profit < 0 :
				hojaExcel.cell(row=nf, column=10).font = Font(color= 'ff0000')
				profit = profit * -1
			hojaExcel.cell(row=nf, column=10, value= profit).number_format = '#,##0.00'	 '''
			nf = nf + 1
		hojaExcel.cell(row=nf, column=12, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=nf, column=13, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=14, value=costTotal).number_format = '#,##0.00'

		hojaExcel.cell(row=nf+2, column=12, value=callTotal+callTotalwr).number_format = '#,##0'
		hojaExcel.cell(row=nf+2, column=13, value=minTotal+minTotalwr).number_format = '#,##0.00'
		hojaExcel.cell(row=nf+2, column=14, value=costTotal+costTotalwr).number_format = '#,##0.00'
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		fila = filaMax + 1
		hojaExcel['J'+str(fila)] = "Total"
		hojaExcel['J'+str(fila)].alignment = Alignment(horizontal="right", vertical="center")
		hojaExcel['J'+str(fila)].font = Font(b=True, color="000000", name='Arial')
		for row in hojaExcel.iter_cols(min_row=fila, max_row=fila, min_col=12, max_col=15):
		    for cell in row:
				if  str(cell).find('.O') == -1 :
					cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
					cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)

		hojaExcel.row_dimensions[fila+1].height = 3.0
		hojaExcel['J'+str(fila+2)] = "General Total"
		hojaExcel['J'+str(fila+2)].alignment = Alignment(horizontal="right", vertical="center")
		hojaExcel['D'+str(fila+2)].font = Font(b=True,color= '044ea4')

	 	for row in hojaExcel.iter_cols(min_row=fila+2, max_row=fila+2, min_col=12, max_col=15):
		    for cell in row:
				if  str(cell).find('.O') == -1 :
					cell.fill = PatternFill(start_color='ffff00',end_color='ffff00', fill_type='solid') # COLOR AMARILLO
					cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
		#*********************************************************************************************************************
		fila = fila + 4
		hojaExcel['C'+str(fila)] = "Retail"
		hojaExcel['C'+str(fila)].font = Font(b=True,color= '000000')

		hojaExcel['D'+str(fila+1)] = "Client"
		hojaExcel['E'+str(fila+1)] = "Calls"
		hojaExcel['F'+str(fila+1)] = "Minutes"
		hojaExcel['G'+str(fila+1)] = "Cost"

		hojaExcel['E'+str(fila+1)].border = Border(top=thin, left=nada, right=nada, bottom=thin)
		hojaExcel['F'+str(fila+1)].border = Border(top=thin, left=nada, right=nada, bottom=thin)
		hojaExcel['G'+str(fila+1)].border = Border(top=thin, left=nada, right=nada, bottom=thin)

	  	for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=4, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
				cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

		n=fila+2
		nf = n #!!!
		fila = fila = fila + 2 -1 + len(Datos['listRetail']) # len('listRetail') CANTIDAD DE REGISTROS EN DICCIONARIO
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row= n, max_row= n, min_col=4, max_col=7):
			    for cell in row:
					cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO DE 'RETAIL'
		callTotal = 0
		minTotal = 0
		costTotal = 0
		Datos['listRetail'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listRetail'])
		for regRetail in Datos['listRetail']:
			callTotal = callTotal + regRetail['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + regRetail['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + regRetail['Q_COST_'+ nombreHoja]
			revTotal = revTotal + regRetail['Q_REV_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=4, value=regRetail['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=5, value=regRetail['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=6, value=regRetail['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value=regRetail['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			nf = nf + 1
		hojaExcel.cell(row=nf, column=4, value='Total').alignment = Alignment(horizontal="right")
		hojaExcel.cell(row=nf, column=5, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=nf, column=6, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=7, value=costTotal).number_format = '#,##0.00'
		for row in hojaExcel.iter_cols(min_row= nf, max_row= nf, min_col=5, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffff00',end_color='ffff00', fill_type='solid') # COLOR AMARILLO
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		fila = fila + 3
		hojaExcel['C'+str(fila)] = "DID"
		hojaExcel['C'+str(fila)].font = Font(b=True,color= '000000')

		hojaExcel['D'+str(fila+1)] = "Client"
		hojaExcel['E'+str(fila+1)] = "Calls"
		hojaExcel['F'+str(fila+1)] = "Minutes"
		hojaExcel['G'+str(fila+1)] = "Cost"

	  	for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=4, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
				cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

		n=fila+2
		nf = n #!!!
		fila = fila = fila + 2 -1 + len(Datos['listDid'])
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row= n, max_row= n, min_col=4, max_col=7):
			    for cell in row:
					cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO DE 'DID'
		tempTupla = (callTotal,minTotal,costTotal)
		callTotal = 0
		minTotal = 0
		costTotal = 0
		Datos['listDid'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listDid'])
		for regDid in Datos['listDid']:
			callTotal = callTotal + regDid['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + regDid['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + regDid['Q_COST_'+ nombreHoja]
			revTotal = revTotal + regDid['Q_REV_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=4, value=regDid['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=5, value=regDid['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=6, value=regDid['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value=regDid['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			nf = nf + 1
		hojaExcel.cell(row=nf, column=4, value='Total').alignment = Alignment(horizontal="right")
		hojaExcel.cell(row=nf, column=5, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=nf, column=6, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=7, value=costTotal).number_format = '#,##0.00'

		for row in hojaExcel.iter_cols(min_row= nf, max_row= nf, min_col=5, max_col=7):
		    for cell in row:
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)

		hojaExcel.cell(row=nf+2, column=5, value=callTotal+tempTupla[0]).number_format = '#,##0'
		hojaExcel.cell(row=nf+2, column=6, value=minTotal+tempTupla[1]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf+2, column=7, value=costTotal+tempTupla[2]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf+2, column=4, value="Client General Totals").font = Font(b=True,color= '000000')
		for row in hojaExcel.iter_cols(min_row= nf+2, max_row= nf+2, min_col=5, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffff00',end_color='ffff00', fill_type='solid')
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		return nf + 4
	#*********************************************************************************************************************
	def ProviderInformationdetail(self,hojaExcel,nombreHoja,filaInicio,Datos,colorFranja):
		fila = filaInicio
		hojaExcel['A'+str(fila)].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['B'+str(fila)] = "►"
		hojaExcel['C'+str(fila)] = "Provider Information detail"
		for row in hojaExcel.iter_cols(min_row=fila, max_row=fila, min_col=2, max_col=16):
		    for cell in row:
		    	cell.fill = PatternFill(start_color=colorFranja,end_color=colorFranja, fill_type='solid') # COLOR AZUL
		    	cell.font = Font(b=True, color="ffffff",size = "10", name='Arial')

		hojaExcel['D'+str(fila+1)] = "Wholesale"
		hojaExcel['D'+str(fila+1)].font = Font(b=True,color= '000000')

		hojaExcel['D'+str(fila+2)] = "Provider"
		hojaExcel['E'+str(fila+2)] = "Calls"
		hojaExcel['F'+str(fila+2)] = "Minutes"
		hojaExcel['G'+str(fila+2)] = "Cost"

	 	for row in hojaExcel.iter_cols(min_row=fila+2, max_row=fila+2, min_col=4, max_col=7):
		    for cell in row:
		    	cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		    	cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		for row in hojaExcel.iter_cols(min_row=fila+2, max_row=fila+2, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

		n= fila +3
		nf = n #!!!
		fila = fila = fila + 3 -1 + len(Datos['listSuppWholesale'])
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row=n, max_row=n, min_col=4, max_col=7):
			    for cell in row:
					cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO DE 'WHOLESALE' EN CUADRO 'Provider Information detail'
		callTotal = 0
		minTotal = 0
		costTotal = 0
		Datos['listSuppWholesale'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listSuppWholesale'])
		for regsw in Datos['listSuppWholesale']:
			callTotal = callTotal + regsw['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + regsw['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + regsw['Q_COST_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=4, value=regsw['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=5, value=regsw['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=6, value=regsw['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value=regsw['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			nf = nf + 1
		hojaExcel.cell(row=nf, column=4, value='Total').alignment = Alignment(horizontal="right")
		hojaExcel.cell(row=nf, column=5, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=nf, column=6, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=7, value=costTotal).number_format = '#,##0.00'

		for row in hojaExcel.iter_cols(min_row= nf, max_row= nf, min_col=5, max_col=7):
		    for cell in row:
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)

		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		fila = fila + 3
		hojaExcel['C'+str(fila)] = "Retail"
		hojaExcel['C'+str(fila)].font = Font(b=True,color= '000000')

		hojaExcel['D'+str(fila+1)] = "Provider"
		hojaExcel['E'+str(fila+1)] = "Calls"
		hojaExcel['F'+str(fila+1)] = "Minutes"
		hojaExcel['G'+str(fila+1)] = "Cost"
	  	for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=4, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
				cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1, min_col=5, max_col=16):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

		n=fila + 2
		nf = n #!!!
		fila = fila = fila + 2 - 1 + len(Datos['listSuppRetail'])
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row= n, max_row= n, min_col=4, max_col=7):
			    for cell in row:
					cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		# BLOQUE DE CODIGO QUE LLENA LAS CELDAS CON LOS VALORES DEL CUADRO DE 'RETAIL'
		tempTupla = (callTotal,minTotal,costTotal)
		callTotal = 0
		minTotal = 0
		costTotal = 0
		Datos['listSuppRetail'] = self.ordenarBy('Q_CALL_'+ nombreHoja,Datos['listSuppRetail'])
		for regsr in Datos['listSuppRetail']:
			callTotal = callTotal + regsr['Q_CALL_'+ nombreHoja]
			minTotal = minTotal + regsr['Q_MIN_'+ nombreHoja]
			costTotal = costTotal + regsr['Q_COST_'+ nombreHoja]
			hojaExcel.cell(row=nf, column=4, value=regsr['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=5, value=regsr['Q_CALL_'+ nombreHoja]).number_format = '#,##0'
			hojaExcel.cell(row=nf, column=6, value=regsr['Q_MIN_'+ nombreHoja]).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value=regsr['Q_COST_'+ nombreHoja]).number_format = '#,##0.00'
			nf = nf + 1
		hojaExcel.cell(row=nf, column=4, value='Total').alignment = Alignment(horizontal="right")
		hojaExcel.cell(row=nf, column=5, value=callTotal).number_format = '#,##0'
		hojaExcel.cell(row=nf, column=6, value=minTotal).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=7, value=costTotal).number_format = '#,##0.00'

		for row in hojaExcel.iter_cols(min_row= nf, max_row= nf, min_col=5, max_col=7):
		    for cell in row:
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)

		hojaExcel.cell(row=nf+2, column=5, value=callTotal+tempTupla[0]).number_format = '#,##0'
		hojaExcel.cell(row=nf+2, column=6, value=minTotal+tempTupla[1]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf+2, column=7, value=costTotal+tempTupla[2]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf+2, column=4, value="Client General Totals").font = Font(b=True,color= '000000')
		for row in hojaExcel.iter_cols(min_row= nf+2, max_row= nf+2, min_col=5, max_col=7):
		    for cell in row:
				cell.fill = PatternFill(start_color='ffff00',end_color='ffff00', fill_type='solid')
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)
		#******************************************************************************************************************
		#******************************************************************************************************************
		#******************************************************************************************************************
		return fila + 5
	#*********************************************************************************************************************
	def BilateralsDestinations(self,hojaExcel,nombreHoja,filaInicio,Datos):
		fila = filaInicio
		hojaExcel['A'+str(fila)].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['B'+str(fila)] = "►"
		hojaExcel['C'+str(fila)] = "Bilaterals Destinations (minutes by provider)"
		for row in hojaExcel.iter_cols(min_row=fila, max_row=fila, min_col=2, max_col=19):
		    for cell in row:
		    	cell.fill = PatternFill(start_color='044ea4',end_color='044ea4', fill_type='solid') # COLOR AZUL
		    	cell.font = Font(b=True, color="ffffff",size = "10", name='Arial')

	 	for row in hojaExcel.iter_cols(min_row=fila+2, max_row=fila+2, min_col=4, max_col=19):
		    for cell in row:
		    	cell.fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		    	cell.border = Border(top=nada, left=nada, right=nada, bottom=thin)

		for row in hojaExcel.iter_cols(min_row=fila+2, max_row=fila+2, min_col=5, max_col=19):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")
		#***********************************************************************************************************
		#***********************************************************************************************************
		#***********************************************************************************************************
		# CREACION DE NUEVO DICICONARIO PARA CONTENER LOS DATOS QUE SE INSERTAR EN LA TABLA DE BILATERALS DETINATIONS
		# 'listNombGrupo' ES UNA LISTA CON LOS NOMBRES DE LOS PAISES DESTINO ORDENADO ALFABETICAMENTE
		listNombGrupo = []
		nombreBuffer = ''
		for i in Datos['listBilatDest']:
			if nombreBuffer.find(i['NB_GRUPO']) == -1:
				listNombGrupo.append(i['NB_GRUPO'])
				nombreBuffer += i['NB_GRUPO']
		listNombGrupo.sort()
		# 'listNombCortoOrden' ES UNA LISTA CON LOS NOMBRES CORTOS DE LAS OPERADORAS ORDENADOS SEGUN LOS INDICADO EN LA BD
		listNombCortoOrden = ['DESTINATIONS']
		for diccNombCortoOrden in Datos['nombCortosOrden']:
			listNombCortoOrden.append(diccNombCortoOrden['NB_CORTO_PROV'])
		while len(listNombCortoOrden) < 11:
			listNombCortoOrden.append('')
		listNombCortoOrden.append('OTHERS')
		listNombCortoOrden.append('TOTAL')
		# 'listNB_PROV' ES UNA LISTA CON LOS NOMBRES COMPLETO DE LAS OPERADORAS
		listNB_PROV = []
		for diccNombCorto in Datos['nombCortos']:
			listNB_PROV.append(diccNombCorto['NB_PROV'])
		#******************************************************************************************************************
		nf = fila + 2 # !!!!
		n= fila + 3
		fila = fila + 3 -1 + len(listNombGrupo)
		while (n<=fila):
		 	for row in hojaExcel.iter_cols(min_row=n, max_row=n, min_col=4, max_col=19):
			    for cell in row:
					cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			n = n + 2
		#******************************************************************************************************************
		diccTemp = {}
		listBD = []
		for paisDest in listNombGrupo:
			# SE CREA EL NUEVO DICCIONARIO PARA BILATERALS DETINATIONS
			diccTemp['DESTINATIONS'] = paisDest
			diccTemp['Q_CALL_'+nombreHoja] = 0
			# Datos['nombCortosOrden'] ES UNA TUPLA DE DICCIONARIOS CON LOS NOMBRE CORTOS ORDENADOS
			# SEGUN DEBEN ESTAR EN EL REPORTE
			for diccNombOrden in Datos['nombCortosOrden']:
				diccTemp[diccNombOrden['NB_CORTO_PROV']] = 0
			diccTemp['OTHERS'] = 0
			diccTemp['TOTAL'] = 0
			# SE BUSCA LAS COINCIDENCIAS EN EL NOMBRE DEL PAIS BARRIENDO LA LISTA DE DICCIONARIOS Datos['listBilatDest']
			for diccBilatDest in Datos['listBilatDest']:
				if paisDest == diccBilatDest['NB_GRUPO']:
					# Datos['nombCortos'] ES UNA TUPLA DE DICCIONARIOS CON LOS NOMBRE CORTOS 'NB_CORTO_PROV'
					# Y SU EQUIVALENTES 'NB_OPERADORA'
					# SE BUSCA LAS COINCIDENCIAS EN EL NOMBRE DEL CORTO DE OPERADORA
					# BARRIENDO LA TUPLA DE DICCIONARIOS Datos['nombCortos']
					if listNB_PROV.count(diccBilatDest['NB_OPERADORA'])>0:
						for diccNombCorto in Datos['nombCortos']:
							if diccBilatDest['NB_OPERADORA'] == diccNombCorto['NB_PROV']:
								# SE BUSCA LAS COINCIDENCIAS EN EL NOMBRE DEL CORTO EN LA
								# TUPLA DE DICCIONARIOS Datos['nombCortos'] Y SE DETERMINA SI EL VALOR
								# diccBilatDest['Q_MIN_'+ nombreHoja] VA A UN CAMPO ESPECIFICO O AL CAMPO 'OTHERS' DEL DICCIONARIO
								if listNombCortoOrden.count(diccNombCorto['NB_CORTO_PROV']) >0:
									for diccNombOrden in Datos['nombCortosOrden']:
										if diccNombOrden['NB_CORTO_PROV'] == diccNombCorto['NB_CORTO_PROV']:
											diccTemp[diccNombOrden['NB_CORTO_PROV']] = diccTemp[diccNombOrden['NB_CORTO_PROV']]+diccBilatDest['Q_MIN_'+ nombreHoja]
											diccTemp['TOTAL'] = diccTemp['TOTAL'] + diccBilatDest['Q_MIN_'+ nombreHoja]
					else:
						diccTemp['OTHERS'] = diccTemp['OTHERS'] + diccBilatDest['Q_MIN_'+ nombreHoja]
						diccTemp['TOTAL'] = diccTemp['TOTAL'] + diccBilatDest['Q_MIN_'+ nombreHoja]
					diccTemp['Q_CALL_'+nombreHoja] = diccTemp['Q_CALL_'+nombreHoja]+diccBilatDest['Q_CALL_'+ nombreHoja]

			listBD.append(diccTemp)
			diccTemp ={}
		listBD = self.ordenarBy('TOTAL',listBD)
		#print listBD
		#***************************************************************************************************************
		# SE INSERTAN LOS NOMBRES CORTOS EN LAS COLUMNAS SEGUN EL ORDER ESPECIFICADO
 	 	hojaExcel['D'+str(nf)] = listNombCortoOrden[0]
		hojaExcel['E'+str(nf)] = listNombCortoOrden[1]
		hojaExcel['F'+str(nf)] = listNombCortoOrden[2]
		hojaExcel['G'+str(nf)] = listNombCortoOrden[3]
		hojaExcel['H'+str(nf)] = listNombCortoOrden[4]
		hojaExcel['J'+str(nf)] = listNombCortoOrden[5]
		hojaExcel['L'+str(nf)] = listNombCortoOrden[6]
		hojaExcel['M'+str(nf)] = listNombCortoOrden[7]
		hojaExcel['N'+str(nf)] = listNombCortoOrden[8]
		hojaExcel['P'+str(nf)] = listNombCortoOrden[9]
		hojaExcel['Q'+str(nf)] = listNombCortoOrden[10]
		hojaExcel['R'+str(nf)] = listNombCortoOrden[11]
		hojaExcel['S'+str(nf)] = listNombCortoOrden[12]
		# SE INSERTAN LOS DATOS EN LA TABLA BILATERAL DESTINATIONS
		nf = nf + 1
		cont = [0]*12
		for diccBD in listBD:
			for i in range(1,13):
				cont[i-1] = cont[i-1] + diccBD.get(listNombCortoOrden[i],0)

			hojaExcel.cell(row=nf, column=4, value= diccBD.get(listNombCortoOrden[0],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=5, value= diccBD.get(listNombCortoOrden[1],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=6, value= diccBD.get(listNombCortoOrden[2],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=7, value= diccBD.get(listNombCortoOrden[3],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=8, value= diccBD.get(listNombCortoOrden[4],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=10, value= diccBD.get(listNombCortoOrden[5],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=12, value= diccBD.get(listNombCortoOrden[6],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=13, value= diccBD.get(listNombCortoOrden[7],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=14, value= diccBD.get(listNombCortoOrden[8],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=16, value= diccBD.get(listNombCortoOrden[9],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=17, value= diccBD.get(listNombCortoOrden[10],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=18, value= diccBD.get(listNombCortoOrden[11],'')).number_format = '#,##0.00'
			hojaExcel.cell(row=nf, column=19, value= diccBD.get(listNombCortoOrden[12],'')).number_format = '#,##0.00'
			nf = nf + 1
		i = 0
		listNombCortoOrden.remove(listNombCortoOrden[0])
		for val in listNombCortoOrden:
			if val =='':
				cont[i] = ''
			i = i+1
		hojaExcel.cell(row=nf, column=5, value= cont[0]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=6, value= cont[1]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=7, value= cont[2]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=8, value= cont[3]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=10, value= cont[4]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=12, value= cont[5]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=13, value= cont[6]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=14, value= cont[7]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=16, value= cont[8]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=17, value= cont[9]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=18, value= cont[10]).number_format = '#,##0.00'
		hojaExcel.cell(row=nf, column=19, value= cont[11]).number_format = '#,##0.00'
		# SE LIMPIAN DE LA MEMORIA YA NO SE NECESITAN
		listNombGrupo = []
		diccTemp = {}
		listBD = []
		listTemp = []
		cont = []
		#***********************************************************************************************************
		#***********************************************************************************************************
		#***********************************************************************************************************
	 	hojaExcel.cell(row=fila+1, column=4, value= 'Total').alignment = Alignment(horizontal="right", vertical="center")
	 	for row in hojaExcel.iter_cols(min_row=fila+1, max_row=fila+1 , min_col=5, max_col=19):
		    for cell in row:
				cell.border = Border(top=thin, left=nada, right=nada, bottom=thin)

		return fila + 3
	#*********************************************************************************************************************
	def ReportDiario(self,hojaExcel,nombreHoja,Datos,imgDir):
		#print nombreHoja
		img = Image(imgDir)
		hojaExcel.add_image(img, 'M2')

		hojaExcel.sheet_properties.tabColor = 'FFFF00'
		hojaExcel.sheet_view.showGridLines = False

		hojaExcel.column_dimensions['A'].width = 1.5
		hojaExcel.column_dimensions['B'].width = 2.0
		hojaExcel.column_dimensions['C'].width = 1.57
		hojaExcel.column_dimensions['D'].width = 36.0
		hojaExcel.column_dimensions['E'].width = 10.71
		hojaExcel.column_dimensions['F'].width = 12.0
		hojaExcel.column_dimensions['G'].width = 12.0
		hojaExcel.column_dimensions['H'].width = 12.0
		hojaExcel.column_dimensions['I'].width = 0.58
		hojaExcel.column_dimensions['J'].width = 12.0
		hojaExcel.column_dimensions['K'].width = 1.0
		hojaExcel.column_dimensions['L'].width = 12.0
		hojaExcel.column_dimensions['M'].width = 12.0
		hojaExcel.column_dimensions['N'].width = 11.57
		hojaExcel.column_dimensions['O'].width = 1.5
		hojaExcel.column_dimensions['P'].width = 12.43
		hojaExcel.column_dimensions['Q'].width = 10.0
		hojaExcel.column_dimensions['R'].width = 11.0
		hojaExcel.column_dimensions['S'].width = 11.0

		hojaExcel['B2'] = "DAILY REPORT"
		hojaExcel['B2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')
		hojaExcel['B3'] = self.localtime
		hojaExcel['B3'].font = Font(color="044ea4", name='Arial')
		#******************************************************Information Resume********************************************
		filaInicio = self.InformationResume(hojaExcel,nombreHoja,Datos,'044ea4')	# COLOR AZUL
		# filaInicio =  AL NUMERO DE LA FILA DONDE INICIARA EL SIGUIENTE CUADRO DE INFORMACION
		#*************************************************Client Information detail***************************************
		filaInicio = self.ClientInformationDetail(hojaExcel,nombreHoja,filaInicio,Datos,'044ea4')
		#************************************************* Provider Information detail **************************************
		filaInicio = self.ProviderInformationdetail(hojaExcel,nombreHoja,filaInicio,Datos,'044ea4')
		#************************************ Bilaterals Destinations (minutes by provider)**********************************
		filaInicio = self.BilateralsDestinations(hojaExcel,nombreHoja,filaInicio,Datos)
	#**********************************************************************************************************************
	def AccumulatedBy(self,hojaExcel,tipoAccumulated,AccumBy,Datos,imgDir):
		img = Image(imgDir)
		hojaExcel.add_image(img, 'K2')

		hojaExcel.sheet_properties.tabColor = '0000ff'
		hojaExcel.sheet_view.showGridLines = False

		hojaExcel.column_dimensions['A'].width = 1.5
		hojaExcel.column_dimensions['B'].width = 35.0
		hojaExcel.column_dimensions['C'].width = 17.0
		hojaExcel.column_dimensions['D'].width = 10.0
		hojaExcel.column_dimensions['E'].width = 10.0
		hojaExcel.column_dimensions['F'].width = 10.0
		hojaExcel.column_dimensions['G'].width = 10.0
		hojaExcel.column_dimensions['H'].width = 10.0
		hojaExcel.column_dimensions['I'].width = 10.0
		hojaExcel.column_dimensions['J'].width = 10.0
		hojaExcel.column_dimensions['K'].width = 10.0
		hojaExcel.column_dimensions['L'].width = 10.0
		hojaExcel.column_dimensions['M'].width = 10.0
		hojaExcel.column_dimensions['N'].width = 10.0
		hojaExcel.column_dimensions['O'].width = 10.0
		hojaExcel.column_dimensions['P'].width = 10.0
		hojaExcel.column_dimensions['Q'].width = 10.0
		hojaExcel.column_dimensions['R'].width = 10.0
		hojaExcel.column_dimensions['S'].width = 10.0

		hojaExcel['B2'] = "DAILY REPORT"
		hojaExcel['B2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')

		hojaExcel['G2'] = tipoAccumulated
		hojaExcel['G2'].font = Font(b=True, color="044ea4", size = "18", name='Arial')
		hojaExcel['G2'].alignment = Alignment(horizontal="center", vertical="center")
		hojaExcel['G3'] = self.localtime
		hojaExcel['G3'].font = Font(color="044ea4", name='Arial')
		hojaExcel['G3'].alignment = Alignment(horizontal="center", vertical="center")

		hojaExcel['A3'].fill = PatternFill(start_color='ffcc00',end_color='ffcc00', fill_type='solid') # COLOR MOSTAZA
		hojaExcel['B3'] = "► Wholesale"
		hojaExcel['C3'].font = Font(b=True, color="000000", size = "12", name='Arial')
		hojaExcel.row_dimensions[4].height = 3.0
		hojaExcel.row_dimensions[5].height = 22.0

		hojaExcel['B5'] = "Client"
		hojaExcel['B5'].alignment = Alignment(horizontal="left", vertical="center")
		hojaExcel['C5'] = AccumBy
		n = 1
		for row in hojaExcel.iter_cols(min_row=5, max_row=5, min_col=4, max_col=34):
		    for cell in row:
		    	cell.value = n
		    	cell.alignment = Alignment(horizontal="center", vertical="center")
		    	n = n +1

	 	for row in hojaExcel.iter_cols(min_row=5, max_row=5, min_col=2, max_col=34):
		    for cell in row:
				cell.fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
		 		cell.font = Font(b=True, color="000000", name='Arial')
		#****************************************************************************************************************
		#****************************************************************************************************************
		#****************************************************************************************************************
		tituloGrafica = ''
		if AccumBy.find('Revenue') != -1:
			tipoAccum = 'Q_REV_'
			tituloGrafica = 'Accumulated Revenue by date'

		else:
			tipoAccum = 'Q_MIN_'
			tituloGrafica = 'Accumulated Minutes by date'

		nf = 6
		listAccumTotal = [0]*32 # DE ESTA FORMA LA LISTA TIENE 32 ELEMETOS CERO
		# SE ORDENA LA LISTA POR EL MAYOR MIN O REV ACUMULADO SEGUN SEA EL CASO
		Datos['listWholesale'] = self.ordenarBy(tipoAccum+'SUM',Datos['listWholesale'])
		for diccWholesale in Datos['listWholesale']:
			hojaExcel.cell(row=nf, column=2, value=diccWholesale['NB_OPERADORA'])
			hojaExcel.cell(row=nf, column=2).alignment = Alignment(horizontal="left", vertical="center")
			hojaExcel.cell(row=nf, column=3, value=diccWholesale[tipoAccum + 'SUM']).number_format = '#,##0.00'
			listAccumTotal[0] = listAccumTotal[0] + diccWholesale[tipoAccum + 'SUM']
			for i in range(1,32):
				hojaExcel.cell(row=nf, column=3+i, value=diccWholesale[tipoAccum + str(i)]).number_format = '#,##0.00'
				listAccumTotal[i] = listAccumTotal[i] + diccWholesale[tipoAccum + str(i)]
			nf = nf +1

		numFila = 5 + len(Datos['listWholesale'])
	 	for row in hojaExcel.iter_cols(min_row=5, max_row=numFila, min_col=2, max_col=34):
		    for cell in row:
				cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

		for i in range(0,32):
			hojaExcel.cell(row=numFila+1, column=3+i, value=listAccumTotal[i]).number_format = '#,##0.00'
			hojaExcel.cell(row=numFila+1, column=3+i).alignment = Alignment(horizontal="right", vertical="center")
			hojaExcel.cell(row=numFila+1, column=3+i).fill = PatternFill(start_color='ccffcc',end_color='ccffcc', fill_type='solid') # COLOR VERDE CLARO
			hojaExcel.cell(row=numFila+1, column=3+i).border = Border(top=nada, left=nada, right=nada, bottom=thin)
			# ESTAS LINEAS COLOCAN LOS VALORES EN UNA SOLA COLUMNA PARA QUE LA FUNCION QUE CREA LA GRAFICA PUEDA ENTENDERLA
			hojaExcel.cell(row=55+i, column=5, value=listAccumTotal[i]).number_format = '#,##0.00'
			hojaExcel.cell(row=55+i, column=5).font = Font(color="ffffff") # Font(color="000000")
		# CREA UNA COLUMNA CON UNA SERIE DEL 1 AL 31 PARA QUE LA FUNCION GRAFICA CREE EL EJE X
		n = 1
		for row in hojaExcel.iter_cols(min_row=55, max_row=85, min_col=4, max_col=4):
		    for cell in row:
		    	cell.value = n
		    	cell.font = Font(color="ffffff") # Font(color="000000")
		    	n = n +1

		#****************************************************************************************************************
		#****************************************************************************************************************
		#****************************************************************************************************************
		Bar = Reference(hojaExcel, min_col= 4, min_row= 55, max_row= 85)
		Data = Reference(hojaExcel, min_col= 5, max_col= 5, min_row= 55 , max_row= 85)
		celdaGrafico ='C'+str(numFila+4)
		self.funcionGrafica(hojaExcel,5,celdaGrafico,tituloGrafica,Data,Bar)
	#**********************************************************************************************************************
	def funcionGrafica(self,hojaExcel,estilo,celdaGrafico,tituloGrafica,Data,Bar):
		chart = BarChart3D()
		chart.width = 17
		chart.auto_axis=False
		chart.x_axis.unit = 1
		chart.y_axis.unit = 1
		chart.dataLabels = DataLabelList()
		chart.type = "col"
		chart.style = estilo
		chart.title = tituloGrafica
		if tituloGrafica =="Accumulated Minutes by date" or tituloGrafica =="Accumulated Revenue by date":
			chart.dataLabels.showVal  = True
			chart.legend = None
			chart.height = 10 # default is 7.5
			chart.width = 25 # default is 15
		chart.y_axis.title = ''
		chart.x_axis.title = ''
		### Bar Chart 3D
		data = Data #Reference(hojaExcel, min_col= min_col, min_row= min_row, max_col= max_col, max_row= max_row)
		Bar = Bar #Reference(hojaExcel, min_col= min_col-1, min_row= min_row + 1, max_row= max_row)
		chart.add_data(data=data, titles_from_data=True)
		chart.set_categories(Bar)
		hojaExcel.add_chart(chart, celdaGrafico)# UBICACION DE LA GRAFICA
	#**********************************************************************************************************************
	def ordenarBy(self,campo,lista):
		listOrdenada=[]
		for i in lista:
			listOrdenada.append([i[campo],i])
		listOrdenada.sort()
		listOrdenada.reverse()
		lista = []
		for i in listOrdenada:
			lista.append(i[1])
		return lista
