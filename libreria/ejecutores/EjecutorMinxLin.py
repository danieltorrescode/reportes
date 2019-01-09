#coding:utf-8
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.chart import BarChart, Series, Reference,BarChart3D
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Border,colors,Fill,Font,GradientFill,NamedStyle, Side,PatternFill,Protection
from datetime import datetime,timedelta
from FuncionesComunes import *
import sys,time,locale,traceback,calendar
sys.path.append("..\\..")
from libreria.conexionBD import *

class Ejecutor(FuncionesComunes):
	def __init__(self,listaParametros,codSolic,rutaArchivo):
			#print "Metodo Constructor del Ejecutor De Minutos Por Línea De Clientes Retail Venezuela"
			try:
				fDesde = listaParametros["Fecha Desde"]
				fHASTA = listaParametros["Fecha Hasta"]

				fDesde = fDesde.split('/')
				fDesde = datetime(int(fDesde[2]),int(fDesde[1]),int(fDesde[0]))
				fDesde1 = fDesde.strftime('%Y%m%d')
				fDesde2 = fDesde.strftime('%Y-%m-%d')

				fHASTA = fHASTA.split('/')
				fHASTA = datetime(int(fHASTA[2]),int(fHASTA[1]),int(fHASTA[0]))
				fHASTA1 = fHASTA.strftime('%Y%m%d')
				fHASTA2 = fHASTA.strftime('%Y-%m-%d')

				fecha = datetime.now().strftime('%Y%m%d%H%M%S')
				yymm = listaParametros["Periodo"]

				query = "SELECT OP.C_OPERADORA,OP.NB_OPERADORA,OP.X_CAMPO_USUARIO1,A.TAS_CASO_TRAFICO_OPERADORA,SUM(A.DURATION),"
				query +="SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT),A.BNO,A.USER_FIELD5,A.TAS_LISTA_PRECIO_QRED_SEG_UNID,"
				query +="A.TAS_LISTA_PRECIO_QRED_MIN_UNID,A.TAS_LISTA_PRECIO_IRED_AJUSTE,M.AB_MONEDA,"
				query +="A.TAS_LISTA_PRECIO_QTASA_MIN,SUM(A.TAS_LISTA_PRECIO_MONTO),count(*) "
				query +="FROM ICX_TRAFICO A "
				query += "INNER JOIN ICX_MONEDA M ON M.C_MONEDA = A.TAS_LISTA_PRECIO_MONEDA "
				query += "INNER JOIN ICX_OPERADORAS OP ON OP.C_TARDEST = "
				query += "IFNULL(SUBSTRING_INDEX(SUBSTRING_INDEX(A.USER_FIELD5,'|',2),'|',-1),'') "
				query += " AND A.C_TIPO_CDR = 'NEXTONE' AND A.TAS_CASO_TRAFICO_DIR_CONTABLE = '2' "
				query += "AND A.F_CDR BETWEEN '"+fDesde2+"' AND '"+ fHASTA2+"' "
				query += "GROUP BY OP.NB_OPERADORA,OP.X_CAMPO_USUARIO1,A.TAS_CASO_TRAFICO_OPERADORA,A.BNO,A.USER_FIELD5,"
				query += "A.TAS_LISTA_PRECIO_QRED_SEG_UNID,A.TAS_LISTA_PRECIO_QRED_MIN_UNID,A.TAS_LISTA_PRECIO_IRED_AJUSTE,"
				query += "M.AB_MONEDA,A.TAS_LISTA_PRECIO_QTASA_MIN,OP.C_OPERADORA"

				#registros = ejecutarQuery(query)
				'''
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print registros
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print len(registros)'''
				libroExcel = Workbook()
				hojaExcel1 = libroExcel.active
				hojaExcel1.title = "Reporte"

				hojaExcel1["A1"] = "REPORTE DE MINUTOS POR LÍNEA DE CLIENTES RETAIL VENEZUELA"
				hojaExcel1["A1"].font = Font(b=True, color="000000", size = 12)

				hojaExcel1.cell(column=1, row= 3).value = "C_CLIENTE"
				hojaExcel1.cell(column=2, row= 3).value = "NB_EMPRESA_CONT_CLIENTE"
				hojaExcel1.cell(column=3, row= 3).value = "ANI"
				hojaExcel1.cell(column=4, row= 3).value = "TIPO_ANI"
				hojaExcel1.cell(column=5, row= 3).value = "Q_CARGOS"
				hojaExcel1.cell(column=6, row= 3).value = "Q_DURACION_REAL"
				hojaExcel1.cell(column=7, row= 3).value = "Q_DURACION_FACTURABLE"
				hojaExcel1.cell(column=8, row= 3).value = "RED (SEG UNIDAD)"
				hojaExcel1.cell(column=9, row= 3).value = "RED (MINIMA UNIDAD)"
				hojaExcel1.cell(column=10, row= 3).value = "RED (UNIDAD ADICIONAL)"
				hojaExcel1.cell(column=11, row= 3).value = "Q_MINUTO_FACTURABLE"
				hojaExcel1.cell(column=12, row= 3).value = "NB_MONEDA"
				hojaExcel1.cell(column=13, row= 3).value = "Q_TARIFA_BASE"
				hojaExcel1.cell(column=14, row= 3).value = "Q_MONTO"
				hojaExcel1.cell(column=15, row= 3).value = "TIPO_ORIGEN"
				hojaExcel1.cell(column=16, row= 3).value = "TIPO_DESTINO"

				# ESTABLECE ESTILOS Y FORMATO A LASS CELDAS

				thin = Side(border_style="thin", color="000000")
				double = Side(border_style="double", color="000000")
				thick = Side(border_style="thick", color="000000")

				border = Border(top=thin, left=thin, right=thin, bottom=thin)
				fill = PatternFill(start_color='4db8ff',end_color='4db8ff', fill_type='solid')
				font = Font(b=True, color="000000")
				alignment = Alignment(horizontal="center", vertical="center")

				def estiloFila(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.font = font
							cell.fill = fill
							cell.border = border

				def alinear(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.border = border

				alinear(3,3,16)
				estiloFila(3,3,16)

				# ESTABLECE ANCHO DE LAS COLUMNAS
				for i in range(17):
				  	hojaExcel1.column_dimensions[get_column_letter(i+1)].width = 27.0

				# ESCRIBIR LOS DATOS
				i = 4
				for reg in ejecutarQuery_v2(50000,query):
					if ('RT_CDTVP','RT_NET1','RT_RESVP','RT_CORVP').count(str(reg["C_OPERADORA"])) == 1:
						pass
					hojaExcel1.cell(column= 1 ,row= i).value = reg["NB_OPERADORA"]
					hojaExcel1.cell(column= 2 ,row= i).value = reg["X_CAMPO_USUARIO1"]
					hojaExcel1.cell(column= 3 ,row= i).value = self.NRO_A(str(reg["USER_FIELD5"]))
					hojaExcel1.cell(column= 4 ,row= i).value = "PENDIENTE"
					hojaExcel1.cell(column= 5 ,row= i).value = reg["count(*)"]
					hojaExcel1.cell(column= 6 ,row= i).value = reg["SUM(A.DURATION)"]
					hojaExcel1.cell(column= 7 ,row= i).value = reg["SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT)"]
					hojaExcel1.cell(column= 8 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_SEG_UNID"]
					hojaExcel1.cell(column= 9 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_MIN_UNID"]
					hojaExcel1.cell(column= 10 ,row= i).value = reg["TAS_LISTA_PRECIO_IRED_AJUSTE"]
					hojaExcel1.cell(column= 11 ,row= i).value = round(reg["SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT)"]/60,6)
					hojaExcel1.cell(column= 12 ,row= i).value = reg["AB_MONEDA"]
					hojaExcel1.cell(column= 13 ,row= i).value = round(reg["TAS_LISTA_PRECIO_QTASA_MIN"],6)
					hojaExcel1.cell(column= 14 ,row= i).value = round(reg["SUM(A.TAS_LISTA_PRECIO_MONTO)"],6)
					hojaExcel1.cell(column= 15 ,row= i).value = self.origen(str(reg["USER_FIELD5"]))
					hojaExcel1.cell(column= 16 ,row= i).value = self.destino(str(reg["BNO"]))

					i=i+1

				alinear(4,3+len(registros),16)

				nombreReporte = str(yymm) + "_ClientesRetailVzla_PorNroA_resumido.xlsx"
				libroExcel.save(rutaArchivo + nombreReporte)

				###############################################################################
				###############################################################################
				###############################################################################
				# REPORTE ClientesRetailVzla detallado
				datosReporte = "REPORTE DE MINUTOS, POR LINEA DE CLIENTES, RETAIL VENEZUELA" + "\r\n"
				nombreReporte2 = str(yymm) +"_ClientesRetailVzla_PorNroA_detallado.csv"
				clienteDetallado = open(rutaArchivo + nombreReporte2,"w")
				clienteDetallado.write(datosReporte)
				nombreCampos = "C_CLIENTE,NB_EMPRESA_CONT_CLIENTE,C_PROVEEDOR,NB_EMPRESA_CONT_PROVEEDOR,ANI,TIPO_ANI,"
				nombreCampos += "C_PREFIJO_GEOGRAFICO,NB_DESTINO,Q_CARGOS,Q_DURACION_REAL,Q_DURACION_FACTURABLE,"
				nombreCampos += "RED (SEG UNIDAD),RED (MINIMA UNIDAD),RED (UNIDAD ADICIONAL),Q_MINUTO_FACTURABLE,"
				nombreCampos += "NB_MONEDA,Q_TARIFA_BASE,Q_MONTO,TIPO_ORIGEN,TIPO_DESTINO"+ "\r\n"
				clienteDetallado.write(nombreCampos)

				query = "SELECT OP.C_OPERADORA,OP.NB_OPERADORA AS 'nombOper1',OP.X_CAMPO_USUARIO1 AS 'campUser1',"
				query += "OC.NB_OPERADORA AS 'nombOper2',OC.X_CAMPO_USUARIO1 AS 'campUser2',A.PREP_BNO_PREFIX,"
				query += "Z.NB_ZONA,COUNT(*),SUM(A.DURATION),SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT),"
				query += "A.TAS_LISTA_PRECIO_QRED_SEG_UNID,A.TAS_LISTA_PRECIO_QRED_MIN_UNID,A.TAS_LISTA_PRECIO_IRED_AJUSTE,"
				query += "M.AB_MONEDA,A.TAS_LISTA_PRECIO_QTASA_MIN,SUM(A.TAS_LISTA_PRECIO_MONTO),A.BNO,A.USER_FIELD5 "
				query += "FROM ICX_TRAFICO A "
				query += "INNER JOIN ICX_ZONAS Z ON Z.C_TIPO_CDR = A.C_TIPO_CDR AND Z.C_ZONA = A.PREP_BNO_ZONA "
				query += "INNER JOIN ICX_MONEDA M ON M.C_MONEDA = A.TAS_LISTA_PRECIO_MONEDA "
				query += "INNER JOIN ICX_OPERADORAS OC ON A.TAS_CASO_TRAFICO_OPERADORA = OC.C_OPERADORA "
				query += "INNER JOIN ICX_OPERADORAS OP ON OP.C_TARDEST = "
				query += "IFNULL(SUBSTRING_INDEX(SUBSTRING_INDEX(A.USER_FIELD5,'|',2),'|',-1),'') "
				query += " AND A.C_TIPO_CDR = 'NEXTONE' AND A.TAS_CASO_TRAFICO_DIR_CONTABLE = '2' "
				query += "AND A.F_CDR BETWEEN '"+fDesde2+"' AND '"+ fHASTA2+"' "
				query += "GROUP BY OP.C_OPERADORA,OP.NB_OPERADORA ,OP.X_CAMPO_USUARIO1,OC.NB_OPERADORA,OC.X_CAMPO_USUARIO1,"
				query += "A.PREP_BNO_PREFIX,Z.NB_ZONA,A.TAS_LISTA_PRECIO_QRED_SEG_UNID,A.TAS_LISTA_PRECIO_QRED_MIN_UNID,"
				query += "A.TAS_LISTA_PRECIO_IRED_AJUSTE,M.AB_MONEDA,A.TAS_LISTA_PRECIO_QTASA_MIN,A.BNO,A.USER_FIELD5"

				#registros = ejecutarQuery(query)
				'''
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print registros
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print len(registros)'''
				for reg in ejecutarQuery_v2(50000,query):
					linea = str(reg["nombOper1"])+ ","
					linea += str(reg["campUser1"])+ ","
					linea += str(reg["nombOper2"])+ ","
					linea += str(reg["campUser2"])+ ","
					linea += self.origen(str(reg["USER_FIELD5"]))+ ","
					linea += "PENDIENTE"+ ","
					linea += str(reg["PREP_BNO_PREFIX"])+ ","
					linea += str(reg["NB_ZONA"])+ ","
					linea += str(reg["COUNT(*)"])+ ","
					linea += str(reg["SUM(A.DURATION)"])+ ","
					linea += str(reg["SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT)"])+ ","
					linea += str(reg["TAS_LISTA_PRECIO_QRED_SEG_UNID"])+ ","
					linea += str(reg["TAS_LISTA_PRECIO_QRED_MIN_UNID"])+ ","
					linea += str(reg["TAS_LISTA_PRECIO_IRED_AJUSTE"])+ ","
					linea += str(round(reg["SUM(A.TAS_LISTA_PRECIO_DURATION_A_FACT)"]/60,6))+ ","
					linea += str(reg["AB_MONEDA"])+ ","
					linea += str(round(reg["TAS_LISTA_PRECIO_QTASA_MIN"],6))+ ","
					linea += str(round(reg["SUM(A.TAS_LISTA_PRECIO_MONTO)"],6))+ ","
					linea += self.origen(str(reg["USER_FIELD5"]))+ ","
					linea += self.destino(str(reg["BNO"])) + "\r\n"

					if ('RT_CDTVP','RT_NET1','RT_RESVP','RT_CORVP').count(str(reg["C_OPERADORA"])) == 1:
						clienteDetallado.write(linea)

				clienteDetallado.close()

				self.rutaArchivo = rutaArchivo + nombreReporte
				nombreReporte = nombreReporte +" y "+nombreReporte2
				self.contenido = listaParametros["cuerpoCorreo"].replace("[/NB_ARCHIVO]",nombreReporte)
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].decode('latin1').encode('utf8')
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/F_DESDE]",listaParametros["Fecha Desde"])
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/F_HASTA]",listaParametros["Fecha Hasta"])
				self.asunto = listaParametros["tituloCorreo"]
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '{0}', F_ULT_ACT = NOW() WHERE C_SOLICITUD = {1}"
				ejecutarQuery(query.format("Ruta del archivo: " + self.rutaArchivo , codSolic))
				
			except Exception as ex:
				detalles = traceback.format_exc()
				observacion = "Excepcion de tipo {0} . Argumentos:\n{1!r}\nDetalles:\n{2} "
				observacion = observacion.format(type(ex).__name__, ex.args,detalles)
				observacion = observacion.replace("'","*")
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '"+ str(observacion)
				query += "',F_ULT_ACT = NOW() WHERE C_SOLICITUD = "+ str(codSolic) +" AND X_OBS IS NULL;"
				respuesta = ejecutarQuery(query)
