#coding:utf-8
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.chart import BarChart, Series, Reference,BarChart3D
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Border,colors,Fill,Font,GradientFill,NamedStyle, Side,PatternFill,Protection
from datetime import datetime
import sys,time,locale,traceback
sys.path.append("..\\..")
from libreria.conexionBD import *

class Ejecutor(object):
	def __init__(self,listaParametros,codSolic,rutaArchivo):
			#print "Metodo Constructor del Ejecutor Tasacion Sumarizado Preliminar"
			try:
				listaParametros["Direccion_Contable"] = listaParametros["Dirección Contable".decode('utf8').encode('latin1')]
				query="SELECT a.*, b.DNIO_SIGNIFICADO AS 'DIR_CONT',c.AB_MONEDA as 'Moneda' "
				query+="FROM ICX_SOLIC_REPORTE a "
				query+="INNER JOIN ICX_DOMINIO b ON b.DNIO_VALOR='"+listaParametros["Direccion_Contable"]+"' and b.DNIO_NOMBRE='DNIO_DIRECCION_CONTABLE' "
				query+="INNER JOIN ICX_MONEDA c ON c.C_MONEDA='"+listaParametros["Moneda"]+"' "
				query+="WHERE a.C_SOLICITUD ="+ str(codSolic)

				SolicitudSumTasaPre = ejecutarQuery(query)
				SolicitudSumTasaPre = SolicitudSumTasaPre[0]
				print SolicitudSumTasaPre

				query = "SELECT b.AB_LISTA_PRECIO, c.AB_PRECIO_DET,a.TAS_LISTA_PRECIO_TPRECIO as T_PRECIO,d.F_INICIO_VIG,"
				query +="a.IA_SERV_CLASS_EXT,a.PREP_ANO_ORIGEN_TASACION,a.PREP_BNO_ZONA,IFNULL(e.NB_ZONA,p.NB_PAIS) NB_ZONA, "
				query +="COUNT(*) AS TOTAL_CANT_CDRS,SUM(a.DURATION) AS TOTAL_DURATION, "
				query +="SUM(a.TAS_LISTA_PRECIO_DURATION_A_FACT) AS TOTAL_DURATION_A_FACT, "
				query +="a.TAS_LISTA_PRECIO_QTASA_MIN,SUM(a.TAS_LISTA_PRECIO_MONTO) AS TOTAL_MONTO, "
				query +="a.TAS_LISTA_PRECIO_QRED_SEG_UNID,a.TAS_LISTA_PRECIO_QRED_MIN_UNID,a.TAS_LISTA_PRECIO_IRED_AJUSTE, "
				query +="CASE WHEN date_format(a.F_CDR,'%Y-%m-%d') between '"+str(listaParametros["Fecha Desde"])+"' "
				query += "and '"+str(listaParametros["Fecha Hasta"])+"' THEN 'N' ELSE 'Y' END AS Derivado, "
				query +="date_format(a.F_CDR,'%Y%m%d') as diaRemanente "
				query +="FROM ICX_TRAFICO a "
				query +="INNER JOIN ICX_LISTA_PRECIO b on (a.C_TIPO_CDR = b.C_TIPO_CDR and a.TAS_LISTA_PRECIO = b.C_LISTA_PRECIO) "
				query +="INNER JOIN ICX_NOMBRE_LISTA_PRECIO c on (a.C_TIPO_CDR = c.C_TIPO_CDR and a.TAS_LISTA_PRECIO = c.C_LISTA_PRECIO "
				query +="and a.TAS_LISTA_PRECIO_DET = c.C_LISTA_PRECIO_DET)  "
				query +="INNER JOIN ICX_NOMBRE_LISTA_PRECIO_DET d on (a.C_TIPO_CDR = d.C_TIPO_CDR and a.TAS_LISTA_PRECIO = d.C_LISTA_PRECIO "
				query +="and a.TAS_LISTA_PRECIO_DET = d.C_LISTA_PRECIO_DET and a.TAS_LISTA_PRECIO_RITEM = d.R_ITEM) "
				query +="LEFT JOIN ICX_ZONAS e on (a.C_TIPO_CDR = e.C_TIPO_CDR and a.PREP_BNO_ZONA = e.C_ZONA) "
				query +="LEFT JOIN ICX_PAIS p ON (a.C_TIPO_CDR = p.C_TIPO_CDR AND a.PREP_BNO_ZONA = p.ABR_PAIS) "
				query +="WHERE a.TAS_CASO_TRAFICO_OPERADORA = '"+str(listaParametros["Operadora"])+"' "
				query +="AND date_format(a.F_CDR,'%Y-%m-%d') between '"+str(listaParametros["Fecha Desde"])+"' and '"+str(listaParametros["Fecha Hasta"])+"' "
				query +="AND a.TAS_CASO_TRAFICO_DIR_CONTABLE = "+str(listaParametros["Direccion_Contable"])+" "
				query +="AND a.TAS_LISTA_PRECIO_MONEDA = '"+str(listaParametros["Moneda"])+"' "
				query +="AND a.C_TIPO_CDR='"+SolicitudSumTasaPre["C_TIPO_CDR"]+"' "
				query +="GROUP BY b.AB_LISTA_PRECIO, c.AB_PRECIO_DET,TAS_LISTA_PRECIO_TPRECIO,d.F_INICIO_VIG, "
				query +="a.IA_SERV_CLASS_EXT,a.PREP_ANO_ORIGEN_TASACION,a.PREP_BNO_ZONA,IFNULL(e.NB_ZONA,p.NB_PAIS), "
				query +="a.TAS_LISTA_PRECIO_QTASA_MIN,a.TAS_LISTA_PRECIO_QRED_SEG_UNID, "
				query +="a.TAS_LISTA_PRECIO_QRED_MIN_UNID,a.TAS_LISTA_PRECIO_IRED_AJUSTE,a.F_CDR ; "

				'''
				print query
				registros = ejecutarQuery(query)
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print registros
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print len(registros)
				'''
				libroExcel = Workbook()
				hojaExcel1 = libroExcel.active

				hojaExcel1.title = "Reporte"

				hojaExcel1["A1"] = "REPORTE DE TASACIÓN PRELIMINAR SUMARIZADO"
				hojaExcel1["A1"].font = Font(b=True, color="000000", size = 12)

				hojaExcel1.cell(column=1, row= 2).value = "Operadora"
				hojaExcel1.cell(column=2, row= 2).value = "Tipo CDR"
				hojaExcel1.cell(column=3, row= 2).value = "Dirección Contable"
				hojaExcel1.cell(column=4, row= 2).value = "Moneda"

				hojaExcel1.cell(column=1, row=3).value = listaParametros["Operadora"]
				hojaExcel1.cell(column=2, row=3).value = SolicitudSumTasaPre["C_TIPO_CDR"]
				hojaExcel1.cell(column=3, row=3).value = SolicitudSumTasaPre["DIR_CONT"]
				hojaExcel1.cell(column=4, row=3).value = SolicitudSumTasaPre["Moneda"]

				hojaExcel1.cell(column=1, row= 5).value = "LISTA PRECIO"
				hojaExcel1.cell(column=2, row= 5).value = "NOMBRE PRECIO"
				hojaExcel1.cell(column=3, row= 5).value = "TIPO PRECIO"
				hojaExcel1.cell(column=4, row= 5).value = "FECHA INICIO VIG"
				hojaExcel1.cell(column=5, row= 5).value = "CLASE SERVICIO"
				hojaExcel1.cell(column=6, row= 5).value = "ORIGEN"
				hojaExcel1.cell(column=7, row= 5).value = "COD ZONA DESTINO"
				hojaExcel1.cell(column=8, row= 5).value = "ZONA DESTINO"
				hojaExcel1.cell(column=9, row= 5).value = "CANTIDAD CARGOS"
				hojaExcel1.cell(column=10, row= 5).value = "DURACION (SEG)"
				hojaExcel1.cell(column=11, row= 5).value = "DURACION FACTURABLE (SEG)"
				hojaExcel1.cell(column=12, row= 5).value = "TARIFA POR MINUTO"
				hojaExcel1.cell(column=13, row= 5).value = "MONTO"
				hojaExcel1.cell(column=14, row= 5).value = "RED (SEG UNIDAD)"
				hojaExcel1.cell(column=15, row= 5).value = "RED (MINIMA UNIDAD)"
				hojaExcel1.cell(column=16, row= 5).value = "RED (UNIDAD ADICIONAL)"
				hojaExcel1.cell(column=17, row= 5).value = "REMANENTE?"
				hojaExcel1.cell(column=18, row= 5).value = "DIA TRAFICO"

				# ESTABLECE ESTILOS Y FORMATO A LASS CELDAS

				thin = Side(border_style="thin", color="000000")
				double = Side(border_style="double", color="000000")
				thick = Side(border_style="thick", color="000000")

				border = Border(top=thin, left=thin, right=thin, bottom=thin)
				fill = PatternFill(start_color='4db8ff',end_color='4db8ff', fill_type='solid')
				font = Font(b=True, color="000000")
				alignment = Alignment(horizontal="center", vertical="center")

				def estiloFila(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.font = font
							cell.fill = fill
							cell.border = border

				def alinear(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.border = border

				estiloFila(2,2,4)
				alinear(3,3,4)
				estiloFila(5,5,18)

				# ESTABLECE ANCHO DE LAS COLUMNAS
				for i in range(19):
				  	hojaExcel1.column_dimensions[get_column_letter(i+1)].width = 27.0

				# ESCRIBIR LOS DATOS
				i = 6
				for reg in ejecutarQuery_v2(1000,query):
					hojaExcel1.cell(column= 1 ,row= i).value = reg["AB_LISTA_PRECIO"]
					hojaExcel1.cell(column= 2 ,row= i).value = reg["AB_PRECIO_DET"]
					hojaExcel1.cell(column= 3 ,row= i).value = reg["T_PRECIO"]
					hojaExcel1.cell(column= 4 ,row= i).value = reg["F_INICIO_VIG"].strftime("%d/%m/%Y")
					hojaExcel1.cell(column= 5 ,row= i).value = reg["IA_SERV_CLASS_EXT"]
					hojaExcel1.cell(column= 6 ,row= i).value = reg["PREP_ANO_ORIGEN_TASACION"]
					hojaExcel1.cell(column= 7 ,row= i).value = reg["PREP_BNO_ZONA"]
					hojaExcel1.cell(column= 8 ,row= i).value = reg["NB_ZONA"]
					hojaExcel1.cell(column= 9 ,row= i).value = reg["TOTAL_CANT_CDRS"]
					hojaExcel1.cell(column= 10 ,row= i).value = reg["TOTAL_DURATION"]
					hojaExcel1.cell(column= 11 ,row= i).value = reg["TOTAL_DURATION_A_FACT"]
					hojaExcel1.cell(column= 12 ,row= i).value = reg["TAS_LISTA_PRECIO_QTASA_MIN"]
					hojaExcel1.cell(column= 13 ,row= i).value = reg["TOTAL_MONTO"]
					hojaExcel1.cell(column= 14 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_SEG_UNID"]
					hojaExcel1.cell(column= 15 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_MIN_UNID"]
					hojaExcel1.cell(column= 16 ,row= i).value = reg["TAS_LISTA_PRECIO_IRED_AJUSTE"]
					hojaExcel1.cell(column= 17 ,row= i).value = reg["Derivado"]
					hojaExcel1.cell(column= 18 ,row= i).value = reg["diaRemanente"]
					i=i+1

				alinear(6,5+i-6,18)
				fecha = datetime.now().strftime('%Y%m%d%H%M%S')
				date_str = listaParametros["Fecha Desde"]
				format_str = '%Y-%m-%d'
				datetime_obj = datetime.strptime(date_str, format_str)
				listaParametros["Fecha Desde"] = datetime_obj.strftime('%Y%m%d')

				date_str = listaParametros["Fecha Hasta"]
				format_str = '%Y-%m-%d'
				datetime_obj = datetime.strptime(date_str, format_str)
				listaParametros["Fecha Hasta"] = datetime_obj.strftime('%Y%m%d')

				nombreReporte = "ICX_"+listaParametros["Operadora"]
				nombreReporte+= "_FDESDE_"+str(listaParametros["Fecha Desde"])+"_FHASTA_"
				nombreReporte+= str(listaParametros["Fecha Hasta"])+ "_" + SolicitudSumTasaPre["DIR_CONT"].replace(" ","_")
				nombreReporte+= "_" + str(fecha) +".xlsx"
				#nombreReporte = nombreReporte.replace(" ","_")
				#print rutaArchivo + nombreReporte

				libroExcel.save(rutaArchivo + nombreReporte)
				self.rutaArchivo = rutaArchivo + nombreReporte
				self.contenido = listaParametros["cuerpoCorreo"].replace("[/NB_ARCHIVO]",nombreReporte)
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].decode('latin1').encode('utf8')
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/C_OPERADORA]",str(listaParametros["Operadora"]))
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/F_DESDE]",str(listaParametros["Fecha Desde"]))
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/F_HASTA]",str(listaParametros["Fecha Hasta"]))
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/C_DIRECCION_CONTABLE]",str(SolicitudSumTasaPre["DIR_CONT"]))
				self.asunto = listaParametros["tituloCorreo"]
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '{0}', F_ULT_ACT = NOW() WHERE C_SOLICITUD = {1}"
				ejecutarQuery(query.format("Ruta del archivo: " + self.rutaArchivo , codSolic))

			except Exception as ex:
				detalles = traceback.format_exc()
				observacion = "Excepcion de tipo {0} . Argumentos:\n{1!r}\nDetalles:\n{2} "
				observacion = observacion.format(type(ex).__name__, ex.args,detalles)
				observacion = observacion.replace("'","*")
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '"+ str(observacion)
				query += "',F_ULT_ACT = NOW() WHERE C_SOLICITUD = "+ str(codSolic) +" AND X_OBS IS NULL;"
				respuesta = ejecutarQuery(query)
