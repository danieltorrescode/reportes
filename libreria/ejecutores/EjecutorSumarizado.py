#coding:utf-8
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.chart import BarChart, Series, Reference,BarChart3D
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Border,colors,Fill,Font,GradientFill,NamedStyle, Side,PatternFill,Protection
from datetime import datetime
import sys,time,locale,traceback
sys.path.append("..\\..")
from libreria.conexionBD import *

class Ejecutor(object):
	def __init__(self,listaParametros,codSolic,rutaArchivo):
			#print "Metodo Constructor del Ejecutor Sumarizado"
			try:
				query = "SELECT A.C_SOLICITUD, A.C_TIPO_CDR, A.C_OPERADORA,A.C_PERIODO,"
				query +="B.DNIO_SIGNIFICADO AS 'C_DIRECCION_CONTABLE',C.AB_MONEDA "
				query += "FROM ICX_SOLIC_CIERRE A "
				query +="INNER JOIN ICX_DOMINIO B ON B.DNIO_VALOR = A.C_DIRECCION_CONTABLE "
				query +="INNER JOIN ICX_MONEDA C ON C.C_MONEDA= A.C_MONEDA "
				query +=" WHERE A.C_SOLICITUD = " + str(listaParametros["Cod Solicitud Cierre"])
				query +=" and B.DNIO_NOMBRE = 'DNIO_DIRECCION_CONTABLE'"

				solicitudCierre = ejecutarQuery(query)

				query = "SELECT E.AB_LISTA_PRECIO, D.AB_PRECIO_DET, C.T_PRECIO,C.F_INICIO_VIG,B.IA_SERV_CLASS_EXT,B.PREP_BNO_ZONA"
				query +=",IFNULL(F.NB_ZONA,P.NB_PAIS) NB_ZONA,SUM(B.TOTAL_CANT_CDRS) AS 'TOTAL_CANT_CDRS',SUM(B.TOTAL_DURATION) AS 'TOTAL_DURATION' "
				query +=",SUM(B.TOTAL_DURATION_A_FACT) AS 'TOTAL_DURATION_A_FACT',B.TAS_LISTA_PRECIO_QTASA_MIN"
				query +=",SUM(B.TOTAL_MONTO) AS 'TOTAL_MONTO',B.TAS_LISTA_PRECIO_QRED_SEG_UNID,B.TAS_LISTA_PRECIO_QRED_MIN_UNID"
				query +=",B.PREP_ANO_ORIGEN_TASACION,B.TAS_LISTA_PRECIO_IRED_AJUSTE,CASE WHEN date_format(B.F_CDR,'%Y-%m-%d') between "
				query +="G.F_INICIO_PERIODO and G.F_FIN_PERIODO THEN 'N' ELSE 'Y' END AS Derivado, date_format(B.F_CDR,'%Y%m%d') as diaRemanente"
				query +=" FROM  ICX_SOLIC_CIERRE A INNER JOIN ICX_CIERRE B ON (A.C_SOLICITUD = B.C_SOLICITUD) "
				query +="INNER JOIN ICX_LISTA_PRECIO E ON (E.C_TIPO_CDR=B.C_TIPO_CDR AND E.C_LISTA_PRECIO = B.TAS_LISTA_PRECIO) "
				query +="INNER JOIN ICX_NOMBRE_LISTA_PRECIO D ON (D.C_TIPO_CDR=B.C_TIPO_CDR AND D.C_LISTA_PRECIO = B.TAS_LISTA_PRECIO "
				query +="AND D.C_LISTA_PRECIO_DET = B.TAS_LISTA_PRECIO_DET) "
				query +="INNER JOIN ICX_NOMBRE_LISTA_PRECIO_DET C ON (C.C_TIPO_CDR=B.C_TIPO_CDR AND C.C_LISTA_PRECIO = B.TAS_LISTA_PRECIO "
				query +="AND C.C_LISTA_PRECIO_DET = B.TAS_LISTA_PRECIO_DET AND C.R_ITEM = B.TAS_LISTA_PRECIO_RITEM) "
				query +="LEFT JOIN ICX_ZONAS F ON (F.C_TIPO_CDR = B.C_TIPO_CDR AND F.C_ZONA = B.PREP_BNO_ZONA) "
				query +="LEFT JOIN ICX_PAIS P ON (P.C_TIPO_CDR = B.C_TIPO_CDR AND P.ABR_PAIS = B.PREP_BNO_ZONA) "
				query +="INNER JOIN ICX_PERIODOS_DET G ON (G.C_TIPO_CDR = A.C_TIPO_CDR AND G.C_PLAN_PERIODO = A.C_PLAN_PERIODO "
				query +="AND G.C_PERIODO = A.C_PERIODO) "
				query +="WHERE B.C_SOLICITUD = " + str(listaParametros["Cod Solicitud Cierre"]) #+ str(codSolic)
				query +=" GROUP BY E.AB_LISTA_PRECIO, D.AB_PRECIO_DET, C.T_PRECIO, C.F_INICIO_VIG,B.IA_SERV_CLASS_EXT,B.PREP_BNO_ZONA,"
				query +="IFNULL(F.NB_ZONA,P.NB_PAIS),B.TAS_LISTA_PRECIO_QTASA_MIN,B.TAS_LISTA_PRECIO_QRED_SEG_UNID,B.TAS_LISTA_PRECIO_QRED_MIN_UNID,"
				query +="B.TAS_LISTA_PRECIO_IRED_AJUSTE,B.F_CDR,B.PREP_ANO_ORIGEN_TASACION"
				#registros = ejecutarQuery(query)
				'''
				print query
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print registros
				print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
				print len(registros)'''
				libroExcel = Workbook()
				hojaExcel1 = libroExcel.active

				hojaExcel1.title = "Reporte"

				hojaExcel1["A1"] = "REPORTE DE TASACIÓN FINAL SUMARIZADO"
				hojaExcel1["A1"].font = Font(b=True, color="000000", size = 12)

				hojaExcel1.cell(column=1, row= 2).value = "Operadora"
				hojaExcel1.cell(column=2, row= 2).value = "Tipo CDR"
				hojaExcel1.cell(column=3, row= 2).value = "Dirección Contable"
				hojaExcel1.cell(column=4, row= 2).value = "Moneda"
				hojaExcel1.cell(column=5, row= 2).value = "Solicitud Cierre"

				hojaExcel1.cell(column=1, row=3).value = solicitudCierre[0]["C_OPERADORA"]
				hojaExcel1.cell(column=2, row=3).value = solicitudCierre[0]["C_TIPO_CDR"]
				hojaExcel1.cell(column=3, row=3).value = solicitudCierre[0]["C_DIRECCION_CONTABLE"]
				hojaExcel1.cell(column=4, row=3).value = solicitudCierre[0]["AB_MONEDA"]
				hojaExcel1.cell(column=5, row=3).value = solicitudCierre[0]["C_SOLICITUD"]

				hojaExcel1.cell(column=1, row= 5).value = "LISTA PRECIO"
				hojaExcel1.cell(column=2, row= 5).value = "NOMBRE PRECIO"
				hojaExcel1.cell(column=3, row= 5).value = "TIPO PRECIO"
				hojaExcel1.cell(column=4, row= 5).value = "FECHA INICIO VIG"
				hojaExcel1.cell(column=5, row= 5).value = "CLASE SERVICIO"
				hojaExcel1.cell(column=6, row= 5).value = "ORIGEN"
				hojaExcel1.cell(column=7, row= 5).value = "COD ZONA DESTINO"
				hojaExcel1.cell(column=8, row= 5).value = "ZONA DESTINO"
				hojaExcel1.cell(column=9, row= 5).value = "CANTIDAD CARGOS"
				hojaExcel1.cell(column=10, row= 5).value = "DURACION (SEG)"
				hojaExcel1.cell(column=11, row= 5).value = "DURACION FACTURABLE (SEG)"
				hojaExcel1.cell(column=12, row= 5).value = "TARIFA POR MINUTO"
				hojaExcel1.cell(column=13, row= 5).value = "MONTO"
				hojaExcel1.cell(column=14, row= 5).value = "RED (SEG UNIDAD)"
				hojaExcel1.cell(column=15, row= 5).value = "RED (MINIMA UNIDAD)"
				hojaExcel1.cell(column=16, row= 5).value = "RED (UNIDAD ADICIONAL)"
				hojaExcel1.cell(column=17, row= 5).value = "REMANENTE?"
				hojaExcel1.cell(column=18, row= 5).value = "DIA TRAFICO"

				# ESTABLECE ESTILOS Y FORMATO A LASS CELDAS
				thin = Side(border_style="thin", color="000000")
				double = Side(border_style="double", color="000000")
				thick = Side(border_style="thick", color="000000")

				border = Border(top=thin, left=thin, right=thin, bottom=thin)
				fill = PatternFill(start_color='4db8ff',end_color='4db8ff', fill_type='solid')
				font = Font(b=True, color="000000")
				alignment = Alignment(horizontal="center", vertical="center")

				def estiloFila(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.font = font
							cell.fill = fill
							cell.border = border

				def alinear(min_row, max_row, max_col):
					for row in hojaExcel1.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
					    for cell in row:
							cell.alignment = alignment
							cell.border = border

				estiloFila(2,2,5)
				alinear(3,3,5)
				estiloFila(5,5,18)
				'''
				for row in hojaExcel1.iter_cols(min_row=12, max_row=20, min_col=18, max_col=18):
				    for cell in row:
				    	cell.font = Font(b=True, color='4db8ff')

				# Define el formato para el separador de miles
				locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252') # English_United States.1252
				'''
				# ESTABLECE ANCHO DE LAS COLUMNAS
				for i in range(19):
				  	hojaExcel1.column_dimensions[get_column_letter(i+1)].width = 27.0

				# ESCRIBIR LOS DATOS
				i = 6
				for reg in ejecutarQuery_v2(50000,query):
					hojaExcel1.cell(column= 1 ,row= i).value = reg["AB_LISTA_PRECIO"]
					hojaExcel1.cell(column= 2 ,row= i).value = reg["AB_PRECIO_DET"]
					hojaExcel1.cell(column= 3 ,row= i).value = reg["T_PRECIO"]
					hojaExcel1.cell(column= 4 ,row= i).value = reg["F_INICIO_VIG"].strftime("%d/%m/%Y")
					hojaExcel1.cell(column= 5 ,row= i).value = reg["IA_SERV_CLASS_EXT"]
					hojaExcel1.cell(column= 6 ,row= i).value = reg["PREP_ANO_ORIGEN_TASACION"]
					hojaExcel1.cell(column= 7 ,row= i).value = reg["PREP_BNO_ZONA"]
					hojaExcel1.cell(column= 8 ,row= i).value = reg["NB_ZONA"]
					hojaExcel1.cell(column= 9 ,row= i).value = reg["TOTAL_CANT_CDRS"]#locale.format("%d", (reg["TOTAL_CANT_CDRS"]), True)
					hojaExcel1.cell(column= 10 ,row= i).value = reg["TOTAL_DURATION"]#locale.format("%d", (reg["TOTAL_DURATION"]), True)
					hojaExcel1.cell(column= 11 ,row= i).value = reg["TOTAL_DURATION_A_FACT"]#locale.format("%d", (reg["TOTAL_DURATION_A_FACT"]), True)
					hojaExcel1.cell(column= 12 ,row= i).value = reg["TAS_LISTA_PRECIO_QTASA_MIN"]#locale.format("%.4f", (reg["TAS_LISTA_PRECIO_QTASA_MIN"]), True)
					hojaExcel1.cell(column= 13 ,row= i).value = reg["TOTAL_MONTO"]#locale.format("%.4f", (reg["TOTAL_MONTO"]), True)
					hojaExcel1.cell(column= 14 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_SEG_UNID"]#locale.format("%d", (reg["TAS_LISTA_PRECIO_QRED_SEG_UNID"]), True)
					hojaExcel1.cell(column= 15 ,row= i).value = reg["TAS_LISTA_PRECIO_QRED_MIN_UNID"]#locale.format("%d", (reg["TAS_LISTA_PRECIO_QRED_MIN_UNID"]), True)
					hojaExcel1.cell(column= 16 ,row= i).value = reg["TAS_LISTA_PRECIO_IRED_AJUSTE"]
					hojaExcel1.cell(column= 17 ,row= i).value = reg["Derivado"]
					hojaExcel1.cell(column= 18 ,row= i).value = reg["diaRemanente"]
					i=i+1

				alinear(6,5+i-6,18)
				fecha = datetime.now().strftime('%Y%m%d%H%M%S')
				nombreReporte = "ICX_"+str(solicitudCierre[0]["C_SOLICITUD"]) +"_"+solicitudCierre[0]["C_OPERADORA"]
				nombreReporte+= "_" + solicitudCierre[0]["C_PERIODO"] + "_" + solicitudCierre[0]["C_DIRECCION_CONTABLE"].replace(" ","_") + "_" + str(fecha) +".xlsx"
				libroExcel.save(rutaArchivo + nombreReporte)
				self.rutaArchivo = rutaArchivo + nombreReporte
				self.contenido = listaParametros["cuerpoCorreo"].replace("[/NB_ARCHIVO]",nombreReporte)
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].decode('latin1').encode('utf8')
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/C_OPERADORA]",str(solicitudCierre[0]["C_OPERADORA"]))
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/C_PERIODO]",str(solicitudCierre[0]["C_PERIODO"]))
				listaParametros["tituloCorreo"] = listaParametros["tituloCorreo"].replace("[/C_DIRECCION_CONTABLE]",str(solicitudCierre[0]["C_DIRECCION_CONTABLE"]))
				self.asunto = listaParametros["tituloCorreo"]
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '{0}', F_ULT_ACT = NOW() WHERE C_SOLICITUD = {1}"
				ejecutarQuery(query.format("Ruta del archivo: " + self.rutaArchivo , codSolic))

			except Exception as ex:
				detalles = traceback.format_exc()
				observacion = "Excepcion de tipo {0} . Argumentos:\n{1!r}\nDetalles:\n{2} "
				observacion = observacion.format(type(ex).__name__, ex.args,detalles)
				observacion = observacion.replace("'","*")
				query = "UPDATE ICX_SOLIC_REPORTE SET X_OBS = '"+ str(observacion)
				query += "',F_ULT_ACT = NOW() WHERE C_SOLICITUD = "+ str(codSolic) +" AND X_OBS IS NULL;"
				respuesta = ejecutarQuery(query)
